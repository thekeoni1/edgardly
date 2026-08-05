"""test_formula_eval.py -- the formulas are evaluated, not just written.

V2_PLAN risk R2 is that formula correctness cannot be verified automatically:
openpyxl writes formula text and computes nothing, so a wiring bug ships
silently and only a human opening Excel sees it. Task 2.5 is the answer. These
tests generate a workbook from a committed fixture, hand it to the `formulas`
library, and read the results back.

Three properties, and they are the ones a hand-check would spend longest on:

  - No cell in the workbook evaluates to an error.
  - The balance check is zero in every historical column, or is the residual
    this filer's own statements produce and the workbook has already explained.
  - With the Assumptions sheet blank every forecast cell is blank, and with it
    filled every forecast cell is a number and all three checks still hold.

This narrows the hand-check rather than replacing it. It cannot tell whether a
row is the right row, whether a tag means what the registry says it means, or
whether a plug is a reasonable size; those are what the acceptance checklist in
V2_PLAN Part 4 is for.

The scaffold's formula vocabulary is deliberately confined to what this library
evaluates -- references, the four operators, IF, OR and ISBLANK -- so a formula
these tests cannot check is one three_statement.py cannot express.

Nothing here touches the network. The `formulas` import is required, not
optional: skipping the harness when the library is absent would turn the one
automated defence against R2 into a test that passes by not running.
"""

import json
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import formulas

from scaffold import excel
from scaffold import three_statement as ts

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")

MILLION = 1_000_000

APPLE = (320193, "3571")
HONEYWELL = (773840, "3728")
KROGER = (56873, "5411")

# One consistent, deliberately unremarkable set. The numbers are not a forecast
# of anything; they exist to make every assumption non-blank so the guards open
# and the wiring can be read. A model that only balances for one set of inputs
# is not balanced, so two of these are negative or unusual on purpose.
FILLED = {
    "rev_growth": 0.05,
    "gross_margin": 0.42,
    "sga_pct_rev": 0.07,
    "rnd_pct_rev": 0.06,
    "da_pct_rev": 0.03,
    "capex_pct_rev": 0.035,
    "sbc_pct_rev": 0.02,
    "dso": 35.0,
    "dio": 22.0,
    "dpo": 60.0,
    "interest_rate": 0.045,
    "tax_rate": 0.21,
    "dividend_payout": 0.20,
    "buyback_pct_ni": 0.55,
    "net_debt_issuance": -2_000 * MILLION,
}

# Balances are in whole dollars and the arithmetic behind them is exact, so a
# residual above a dollar is a real difference. Excel's own arithmetic is
# binary floating point, which on a 400 billion dollar balance sheet leaves
# noise in the fourth decimal place.
TOLERANCE = 1.0


def _facts(cik):
    with open(os.path.join(FIXTURES, "cik{}.json".format(cik)), encoding="utf-8") as h:
        return json.load(h)


def _spec(company, **kwargs):
    cik, sic = company
    return ts.build_model(cik, _facts(cik), sic, **kwargs)


def _fill_assumptions(path, filled_path, values):
    """Type the analyst's numbers into the named input cells and save a copy."""
    workbook = openpyxl.load_workbook(path)
    for name, defined in workbook.defined_names.items():
        if not name.startswith("asm_") or "_ready_" in name:
            continue
        key = name[len("asm_"):].rsplit("_y", 1)[0]
        for title, coordinate in defined.destinations:
            workbook[title][coordinate] = values[key]
    workbook.save(filled_path)
    return filled_path


def _evaluate(path):
    """Every cell of a workbook, keyed "SHEET NAME!A1" as the library upper-cases it."""
    model = formulas.ExcelModel().loads(path).finish()
    solution = model.calculate()
    values = {}
    for key, node in solution.items():
        text = str(key)
        if "!" not in text:
            continue
        try:
            values[text.split("]")[-1].replace("'", "")] = node.value[0, 0]
        except (AttributeError, IndexError, TypeError):
            continue
    return values


def _is_error(value):
    return isinstance(value, str) and value.startswith("#")


def _is_blank(value):
    """The library returns an empty string for our guard and a sentinel for a
    cell that was never written. Both are blank to a reader."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return repr(value) == "empty"


def _row_of(worksheet, label):
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=excel.DEFAULT_LAYOUT.label_column
                          ).value == label:
            return row
    raise AssertionError("no row labelled {!r}".format(label))


def _columns(worksheet):
    """Period label to column letter, read off the sheet rather than assumed."""
    layout = excel.DEFAULT_LAYOUT
    found = {}
    for column in range(layout.first_data_column, worksheet.max_column + 1):
        label = worksheet.cell(row=layout.header_row, column=column).value
        if label:
            found[label] = worksheet.cell(row=layout.header_row,
                                          column=column).column_letter
    return found


class Evaluated(object):
    """A generated workbook, its evaluated values, and the spec behind it."""

    def __init__(self, spec, path, values):
        self.spec = spec
        self.workbook = openpyxl.load_workbook(path)
        self.values = values

    def at(self, sheet, label, period_label):
        worksheet = self.workbook[sheet]
        column = _columns(worksheet)[period_label]
        return self.values.get("{}!{}{}".format(sheet.upper(), column,
                                                _row_of(worksheet, label)))

    def check(self, name_prefix, period_label):
        worksheet = self.workbook["Checks"]
        label = [c["name"] for c in self.spec.checks
                 if c["name"].startswith(name_prefix)][0]
        column = _columns(worksheet)[period_label]
        return self.values.get("CHECKS!{}{}".format(column, _row_of(worksheet, label)))

    def errors(self):
        return {key: value for key, value in self.values.items()
                if _is_error(value)}

    def forecast_cells(self):
        """Every written cell in a forecast column of the three statements."""
        found = {}
        forecast = [p.label for p in ts.forecast_periods(self.spec)]
        for sheet in ("Income Statement", "Balance Sheet", "Cash Flow"):
            worksheet = self.workbook[sheet]
            columns = _columns(worksheet)
            for label in forecast:
                column = columns[label]
                for row in range(excel.DEFAULT_LAYOUT.first_data_row,
                                 worksheet.max_row + 1):
                    coordinate = "{}{}".format(column, row)
                    if worksheet[coordinate].value is None:
                        continue
                    found["{}!{}".format(sheet.upper(), coordinate)] = \
                        self.values.get("{}!{}".format(sheet.upper(), coordinate))
        return found


def _evaluated(company, tmp_path_factory, values=None):
    spec = _spec(company)
    directory = tmp_path_factory.mktemp("eval{}".format(company[0]))
    path = str(directory / "scaffold.xlsx")
    excel.write_workbook(spec, path)
    if values is not None:
        path = _fill_assumptions(path, str(directory / "filled.xlsx"), values)
    return Evaluated(spec, path, _evaluate(path))


@pytest.fixture(scope="module")
def apple_blank(tmp_path_factory):
    return _evaluated(APPLE, tmp_path_factory)


@pytest.fixture(scope="module")
def apple_filled(tmp_path_factory):
    return _evaluated(APPLE, tmp_path_factory, FILLED)


@pytest.fixture(scope="module")
def honeywell_blank(tmp_path_factory):
    return _evaluated(HONEYWELL, tmp_path_factory)


@pytest.fixture(scope="module")
def honeywell_filled(tmp_path_factory):
    return _evaluated(HONEYWELL, tmp_path_factory, FILLED)


@pytest.fixture(scope="module")
def kroger_filled(tmp_path_factory):
    return _evaluated(KROGER, tmp_path_factory, FILLED)


# ---------------------------------------------------------------------------
# Nothing errors
# ---------------------------------------------------------------------------

@pytest.mark.timeout(180)
def test_no_cell_evaluates_to_an_error(apple_blank, honeywell_blank, apple_filled,
                                       honeywell_filled, kroger_filled):
    """A #REF! or a #VALUE! is the wiring bug R2 says ships silently."""
    for book in (apple_blank, honeywell_blank, apple_filled, honeywell_filled,
                 kroger_filled):
        assert book.errors() == {}, book.spec.entity


# ---------------------------------------------------------------------------
# The historical statements, computed
# ---------------------------------------------------------------------------

@pytest.mark.timeout(180)
def test_the_balance_check_is_zero_in_every_historical_column(apple_blank,
                                                              kroger_filled):
    """The exit criterion, evaluated rather than asserted about the spec.

    Apple and Kroger tag all three totals, so this is a real check on real
    numbers: assets, liabilities and equity as three separate tags from three
    separate elements, subtracted in Excel.
    """
    for book in (apple_blank, kroger_filled):
        for period in ts.historical_periods(book.spec):
            value = book.check("Balance check", period.label)
            assert abs(value) < TOLERANCE, "{} {}".format(book.spec.entity,
                                                          period.label)


@pytest.mark.timeout(180)
def test_honeywells_balance_check_is_zero_because_it_was_made_zero(honeywell_blank):
    """The filer with by-design blanks, and the caveat travels with the number.

    Honeywell tags no Liabilities element, so the workbook derives the total
    from assets less equity and the check cannot be Assets = Liabilities +
    Equity verbatim. It evaluates to zero, and the row is flagged and left
    uncoloured so nobody reads the zero as evidence.
    """
    balance = [c for c in honeywell_blank.spec.checks
               if c["name"].startswith("Balance")][0]
    assert balance["tie"] is False
    assert [f["flag_type"] for f in balance["flags"]] == [ts.FLAG_CHECK_NOT_AVAILABLE]

    for period in ts.historical_periods(honeywell_blank.spec):
        assert abs(honeywell_blank.check("Balance check", period.label)) < TOLERANCE


@pytest.mark.timeout(180)
def test_the_cash_tie_is_zero_where_it_can_be_and_documented_where_it_is_not(
        apple_blank, honeywell_blank):
    """Apple's FY2025 closes exactly; Honeywell's residual is its currency effect.

    A filer with foreign operations reports an effect of exchange rates on cash
    that no registry item reads, so the cash flow statement closes short of the
    balance sheet by that amount. Documented here as the filer's own residual,
    which is what the exit criterion allows for.
    """
    assert abs(apple_blank.check("Cash tie", "FY2025")) < TOLERANCE
    assert honeywell_blank.check("Cash tie", "FY2025") == pytest.approx(
        -837 * MILLION, abs=TOLERANCE)


@pytest.mark.timeout(180)
def test_the_workbook_computes_the_same_numbers_the_model_spec_holds(apple_blank,
                                                                    honeywell_blank):
    """Every derived historical cell, evaluated in Excel, against the spec.

    This is the one test that closes the loop R2 opens. three_statement.py
    computes a value in Python and excel.py writes arithmetic it believes
    produces the same value; nothing until now has made the two meet.
    """
    compared = 0
    for book in (apple_blank, honeywell_blank):
        for code, sheet_name, _block in ts.STATEMENT_BLOCKS:
            for row in ts.rows_for(book.spec, code):
                for period in ts.historical_periods(book.spec):
                    cell = row.cells[period.key]
                    if cell.state != ts.CELL_DERIVED:
                        continue
                    evaluated = book.at(sheet_name, row.label, period.label)
                    assert evaluated == pytest.approx(cell.value, abs=TOLERANCE), (
                        "{} {} {}".format(book.spec.entity, row.name, period.label))
                    compared += 1
    # A floor, so a writer that stopped emitting formulas would fail here rather
    # than pass by having nothing to compare. Two filers contribute 184 derived
    # historical cells between them.
    assert compared > 150


@pytest.mark.timeout(180)
def test_the_coverage_rows_evaluate_to_the_share_the_spec_holds(apple_blank,
                                                                honeywell_blank):
    """Open question 11's replacement for the plug warning, computed in Excel.

    The row is written as arithmetic on two cells of the balance sheet rather
    than as a number, so this is where the arithmetic and the spec's own figure
    are made to meet. Apple's equity coverage is negative and Honeywell's is
    over three, which is why neither is clamped.
    """
    compared = 0
    for book in (apple_blank, honeywell_blank):
        sheet = book.workbook["Checks"]
        columns = _columns(sheet)
        for entry in book.spec.coverage:
            row = _row_of(sheet, entry["total_row"])
            for period in ts.historical_periods(book.spec):
                expected = entry["cells"][period.key]
                value = book.values.get("CHECKS!{}{}".format(
                    columns[period.label], row))
                if expected is None:
                    assert _is_blank(value)
                    continue
                assert value == pytest.approx(expected, abs=1e-9), (
                    "{} {} {}".format(book.spec.entity, entry["total_row"],
                                      period.label))
                compared += 1
    assert compared == 50


@pytest.mark.timeout(180)
def test_a_schedule_foots_in_every_historical_column(apple_blank):
    """Opening, movements, residual, closing. The residual is what makes it foot."""
    sheet = apple_blank.workbook["Schedules"]
    columns = _columns(sheet)
    for label in ("FY2022", "FY2023", "FY2024", "FY2025"):
        column = columns[label]

        def value(name):
            return apple_blank.values.get("SCHEDULES!{}{}".format(
                column, _row_of(sheet, name)))

        total = (value("Opening balance") + value("Plus capital expenditure")
                 + value("Less depreciation and amortisation")
                 + value("Impairments, disposals, acquisitions and currency (implied)"))
        assert total == pytest.approx(value("Closing balance"), abs=TOLERANCE), label


# ---------------------------------------------------------------------------
# Blank assumptions, blank forecasts
# ---------------------------------------------------------------------------

@pytest.mark.timeout(180)
def test_a_blank_assumptions_sheet_produces_no_forecast_at_all(apple_blank,
                                                               honeywell_blank):
    """No zeros pretending to be forecasts, which is V2_PLAN's own phrase."""
    for book in (apple_blank, honeywell_blank):
        cells = book.forecast_cells()
        assert cells, book.spec.entity
        for coordinate, value in cells.items():
            assert _is_blank(value), "{} {} = {!r}".format(book.spec.entity,
                                                           coordinate, value)


@pytest.mark.timeout(180)
def test_the_readiness_row_is_false_while_the_inputs_are_blank(apple_blank):
    sheet = apple_blank.workbook["Assumptions"]
    row = _row_of(sheet, "Year ready to model?")
    for column in _columns(sheet).values():
        assert apple_blank.values["ASSUMPTIONS!{}{}".format(column, row)] is False


@pytest.mark.timeout(180)
def test_the_checks_are_blank_in_the_forecast_columns_too(apple_blank):
    """A check on nothing is not a passing check, so it must not read as one."""
    for label in ("FY2026E", "FY2027E", "FY2028E"):
        assert _is_blank(apple_blank.check("Balance check", label))
        assert _is_blank(apple_blank.check("Cash tie", label))


# ---------------------------------------------------------------------------
# Filled assumptions, a model that ties
# ---------------------------------------------------------------------------

@pytest.mark.timeout(180)
def test_filled_assumptions_populate_all_three_statements(apple_filled):
    """Revenue compounds, cost follows the margin, and the bottom line arrives."""
    assert apple_filled.at("Income Statement", "Revenue", "FY2026E") == pytest.approx(
        416_161 * MILLION * 1.05, abs=TOLERANCE)
    assert apple_filled.at("Income Statement", "Gross Profit",
                           "FY2026E") == pytest.approx(
        416_161 * MILLION * 1.05 * 0.42, abs=TOLERANCE)

    for sheet, label in (("Income Statement", "Net Income"),
                         ("Balance Sheet", "Total Assets"),
                         ("Balance Sheet", "Total Equity"),
                         ("Cash Flow", "Cash from Operations"),
                         ("Cash Flow", "Cash, end of period")):
        for period in ("FY2026E", "FY2027E", "FY2028E"):
            value = apple_filled.at(sheet, label, period)
            assert isinstance(value, float), "{} {} {}".format(sheet, label, period)


@pytest.mark.timeout(180)
def test_the_forecast_balance_sheet_ties_for_every_filer(apple_filled,
                                                         honeywell_filled,
                                                         kroger_filled):
    """The property the whole forecast design exists to have.

    Cash comes off the cash flow statement, the working capital line is the
    movement in the balance sheet rows the days assumptions drive, and the one
    plug that is not held flat moves by stock compensation less buybacks. Get
    any of those wrong and this is the test that says so; nothing about the
    historical columns would have noticed.
    """
    for book in (apple_filled, honeywell_filled, kroger_filled):
        for period in ts.forecast_periods(book.spec):
            assert abs(book.check("Balance check", period.label)) < TOLERANCE, (
                "{} {}".format(book.spec.entity, period.label))


@pytest.mark.timeout(180)
def test_the_cash_tie_and_retained_earnings_are_exact_in_the_forecast(apple_filled,
                                                                     kroger_filled):
    """Both are definitions of their rows once the model is doing the arithmetic."""
    for book in (apple_filled, kroger_filled):
        for period in ts.forecast_periods(book.spec):
            assert abs(book.check("Cash tie", period.label)) < TOLERANCE
            assert abs(book.check("Retained earnings", period.label)) < TOLERANCE


@pytest.mark.timeout(180)
def test_the_working_capital_line_is_the_movement_the_balance_sheet_shows(
        apple_filled):
    """The identity that makes the two statements agree, read off both of them."""
    for period in ("FY2027E", "FY2028E"):
        prior = "FY{}E".format(int(period[2:6]) - 1)
        movement = 0.0
        for label, sign in (("Accounts Payable", 1), ("Accounts Receivable", -1),
                            ("Inventory", -1)):
            movement += sign * (apple_filled.at("Balance Sheet", label, period)
                                - apple_filled.at("Balance Sheet", label, prior))
        working_capital = apple_filled.at(
            "Cash Flow",
            "Working capital and other operating items (plug to reported total)",
            period)
        assert working_capital == pytest.approx(movement, abs=TOLERANCE)


@pytest.mark.timeout(180)
def test_clearing_the_assumptions_empties_the_forecast_again(tmp_path):
    """The acceptance checklist step, run as a test: fill them, then delete them.

    A workbook that empties only because it was never filled is not the same
    property. Stale results left behind by a cleared assumption are exactly the
    failure V2_PLAN's "no leftovers" phrase is about.
    """
    spec = _spec(APPLE)
    path = str(tmp_path / "scaffold.xlsx")
    excel.write_workbook(spec, path)
    filled = _fill_assumptions(path, str(tmp_path / "filled.xlsx"), FILLED)

    workbook = openpyxl.load_workbook(filled)
    for name, defined in workbook.defined_names.items():
        if not name.startswith("asm_") or "_ready_" in name:
            continue
        for title, coordinate in defined.destinations:
            workbook[title][coordinate] = None
    cleared = str(tmp_path / "cleared.xlsx")
    workbook.save(cleared)

    book = Evaluated(spec, cleared, _evaluate(cleared))
    assert book.errors() == {}
    for coordinate, value in book.forecast_cells().items():
        assert _is_blank(value), "{} = {!r}".format(coordinate, value)


# ---------------------------------------------------------------------------
# The vocabulary stays inside what the harness can check
# ---------------------------------------------------------------------------

def test_the_scaffold_writes_no_function_this_harness_cannot_evaluate():
    """R2's residual, made a test rather than a hope.

    The library covers only part of Excel, so the plan constrains the
    scaffold's formula vocabulary to what it evaluates. Reading the workbook
    for function calls is what stops that constraint from quietly lapsing the
    next time somebody wants a SUMIF.
    """
    allowed = {"IF", "OR", "ISBLANK", "AND", "ABS", "ISNUMBER"}
    spec = _spec(APPLE)
    workbook = excel.build_workbook(spec)
    seen = set()

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                token = ""
                for character in cell.value:
                    if character.isalpha():
                        token += character
                    else:
                        if character == "(" and token:
                            seen.add(token.upper())
                        token = ""
        for rules in worksheet.conditional_formatting:
            for rule in rules.rules:
                for formula in rule.formula or ():
                    for name in ("AND", "ABS", "ISNUMBER"):
                        if name + "(" in formula:
                            seen.add(name)

    assert seen <= allowed, seen - allowed
