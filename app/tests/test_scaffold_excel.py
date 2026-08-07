"""test_scaffold_excel.py -- the workbook, re-read rather than trusted.

app/scaffold/excel.py writes a file, and openpyxl computes nothing, so the only
thing these tests can check is what is actually in the cells: which are static
numbers, which are formulas, what those formulas say, which named ranges exist,
and whether the file reads back at all. Whether the formulas are right is
task 2.5's job and lives in test_formula_eval.py.

Reading the workbook back on every export is what V2_PLAN risk R4 asks for.
Defined names, comments and conditional formatting are the three openpyxl
features that can trigger Excel's repair dialog, and all three are here.

No test in this module touches the network.
"""

import json
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from scaffold import excel
from scaffold import three_statement as ts

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")

MILLION = 1_000_000

APPLE = (320193, "3571")
HONEYWELL = (773840, "3728")
KROGER = (56873, "5411")


def _facts(cik):
    with open(os.path.join(FIXTURES, "cik{}.json".format(cik)), encoding="utf-8") as h:
        return json.load(h)


def _spec(company, **kwargs):
    cik, sic = company
    return ts.build_model(cik, _facts(cik), sic, **kwargs)


def _reread(spec, tmp_path, name="scaffold.xlsx", layout=excel.DEFAULT_LAYOUT):
    """Write the workbook and load it back, which is the whole point of R4."""
    path = str(tmp_path / name)
    excel.write_workbook(spec, path, layout)
    return openpyxl.load_workbook(path), path


@pytest.fixture(scope="module")
def apple_spec():
    return _spec(APPLE)


@pytest.fixture(scope="module")
def apple_book(apple_spec, tmp_path_factory):
    path = str(tmp_path_factory.mktemp("apple") / "apple.xlsx")
    excel.write_workbook(apple_spec, path)
    return openpyxl.load_workbook(path)


@pytest.fixture(scope="module")
def honeywell_book(tmp_path_factory):
    path = str(tmp_path_factory.mktemp("honeywell") / "honeywell.xlsx")
    excel.write_workbook(_spec(HONEYWELL), path)
    return openpyxl.load_workbook(path)


def _find_row(worksheet, label, layout=excel.DEFAULT_LAYOUT):
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=layout.label_column).value == label:
            return row
    raise AssertionError("no row labelled {!r} on {}".format(label, worksheet.title))


def _column_for(worksheet, period_label, layout=excel.DEFAULT_LAYOUT):
    for column in range(layout.first_data_column, worksheet.max_column + 1):
        if worksheet.cell(row=layout.header_row, column=column).value == period_label:
            return column
    raise AssertionError("no column for {}".format(period_label))


def _cell(worksheet, label, period_label):
    return worksheet.cell(row=_find_row(worksheet, label),
                          column=_column_for(worksheet, period_label))


# ---------------------------------------------------------------------------
# The file exists and reads back
# ---------------------------------------------------------------------------

def test_the_workbook_has_the_seven_sheets_the_plan_names(apple_book):
    assert apple_book.sheetnames == ["Assumptions", "Income Statement",
                                     "Balance Sheet", "Cash Flow", "Schedules",
                                     "Checks", "Source Tags"]


def test_every_acceptance_filer_writes_and_reads_back(tmp_path):
    """R4, on all three. A file openpyxl cannot reopen is one Excel will offer
    to repair, and repairing strips content."""
    for company in (APPLE, HONEYWELL, KROGER):
        workbook, path = _reread(_spec(company), tmp_path,
                                 "cik{}.xlsx".format(company[0]))
        assert len(workbook.sheetnames) == 7
        assert os.path.getsize(path) > 10_000


def test_a_refused_filer_gets_no_workbook_at_all(tmp_path):
    """The refusal is the answer; there is no half-built file to pick up later."""
    spec = ts.build_model(19617, _facts(19617), "6021")

    with pytest.raises(ValueError) as raised:
        excel.write_workbook(spec, str(tmp_path / "bank.xlsx"))
    assert "three-statement template" in str(raised.value)
    assert not os.path.exists(str(tmp_path / "bank.xlsx"))


# ---------------------------------------------------------------------------
# Reported values are static, everything else is a formula
# ---------------------------------------------------------------------------

def test_a_reported_value_is_written_as_the_number_the_filer_tagged(apple_book):
    sheet = apple_book["Income Statement"]
    cell = _cell(sheet, "Revenue", "FY2023")

    assert cell.value == 383_285 * MILLION
    assert cell.font.color.rgb.endswith(excel.COLOR_REPORTED)


def test_a_derived_value_is_written_as_live_arithmetic_not_a_number(apple_book):
    """V2_PLAN asks for this by name: the workbook shows its own math."""
    sheet = apple_book["Balance Sheet"]
    short_term = _cell(sheet, "Short-Term Debt", "FY2023")

    assert isinstance(short_term.value, str) and short_term.value.startswith("=")
    assert short_term.font.color.rgb.endswith(excel.COLOR_DERIVED)

    maturities_row = _find_row(sheet, "Current Maturities of Long-Term Debt")
    paper_row = _find_row(sheet, "Commercial Paper")
    column = sheet.cell(row=maturities_row,
                        column=_column_for(sheet, "FY2023")).column_letter
    assert short_term.value == "={0}{1}+{0}{2}".format(column, maturities_row,
                                                       paper_row)


def test_a_nested_derivation_references_the_derived_cell_not_its_inputs(apple_book):
    """Total Debt points at the Short-Term Debt formula, which points at tags.

    Two levels of formula in the workbook itself, so every leaf a reader can
    reach by following references is a blue cell with a filing behind it. This
    is what "every leaf traces to a tag and filing" means once it is a file
    rather than a data structure.
    """
    sheet = apple_book["Balance Sheet"]
    total = _cell(sheet, "Total Debt", "FY2023")
    short_term_row = _find_row(sheet, "Short-Term Debt")
    long_term_row = _find_row(sheet, "Long-Term Debt")
    column = total.column_letter

    assert total.value == "={0}{1}+{0}{2}".format(column, short_term_row,
                                                  long_term_row)
    assert sheet.cell(row=short_term_row, column=total.column).value.startswith("=")


def test_a_plug_is_the_total_less_its_components_as_a_formula(apple_book):
    sheet = apple_book["Balance Sheet"]
    plug = _cell(sheet, "Other current assets (plug to reported total)", "FY2023")
    column = plug.column_letter
    total_row = _find_row(sheet, "Total Current Assets")

    assert plug.value.startswith("={}{}-".format(column, total_row))
    for component in ("Cash and Equivalents", "Short-Term Investments",
                      "Accounts Receivable", "Inventory"):
        assert "{}{}".format(column, _find_row(sheet, component)) in plug.value


def test_a_missing_value_is_an_empty_cell_and_never_a_formula(apple_book):
    """A formula here would read the hole as a zero and look like an answer.

    Apple carries no goodwill, so gross profit is not the only row that would
    go wrong: any arithmetic over an empty cell silently treats it as nothing.
    The cell is left empty, styled as missing, and its comment says which
    filing to open.
    """
    sheet = apple_book["Balance Sheet"]
    cell = _cell(sheet, "Goodwill", "FY2023")

    assert cell.value is None
    assert cell.font.color.rgb.endswith(excel.COLOR_MISSING)
    assert cell.font.italic
    assert "Not available" in cell.comment.text


def test_a_derived_row_with_a_missing_input_gets_no_formula_either(honeywell_book):
    """Honeywell reports no operating income for FY2021, so EBITDA is blank."""
    sheet = honeywell_book["Income Statement"]
    cell = _cell(sheet, "EBITDA", "FY2021")

    assert cell.value is None
    assert "Operating Income" in cell.comment.text


# ---------------------------------------------------------------------------
# Provenance comments
# ---------------------------------------------------------------------------

def test_a_reported_cell_carries_its_tag_filing_and_accession(apple_book):
    cell = _cell(apple_book["Income Statement"], "Revenue", "FY2023")
    text = cell.comment.text

    assert "Reported by the filer" in text
    assert "RevenueFromContractWithCustomerExcludingAssessedTax" in text
    assert "0000320193" in text


def test_a_derived_cell_carries_its_formula_and_descends_into_nested_inputs(apple_book):
    cell = _cell(apple_book["Balance Sheet"], "Total Debt", "FY2023")
    text = cell.comment.text

    assert "Short-Term Debt + Long-Term Debt" in text
    assert "itself computed as" in text
    assert "LongTermDebtCurrent" in text
    assert "CommercialPaper" in text


def test_a_flagged_cell_carries_the_flag_in_its_comment_and_a_fill(apple_book):
    sheet = apple_book["Cash Flow"]
    cell = _cell(sheet, "Other investing activities (plug to reported total)",
                 "FY2025")

    assert ts.FLAG_PLUG_TOO_LARGE in cell.comment.text
    assert cell.fill.fgColor.rgb.endswith(excel.FILL_FLAG)


def test_a_balance_sheet_plug_is_not_flagged_for_its_size(apple_book):
    """Open question 11: the balance sheet reports coverage instead of warning."""
    sheet = apple_book["Balance Sheet"]
    cell = _cell(sheet, "Other current assets (plug to reported total)", "FY2025")

    assert ts.FLAG_PLUG_TOO_LARGE not in (cell.comment.text if cell.comment else "")


def test_a_row_label_carries_the_registry_note_and_the_forecast_convention(apple_book):
    sheet = apple_book["Balance Sheet"]
    label = sheet.cell(row=_find_row(sheet, "Short-Term Investments"),
                       column=excel.DEFAULT_LAYOUT.label_column)

    assert "Forecast: Held at the last reported balance" in label.comment.text


# ---------------------------------------------------------------------------
# Named inputs and the forecast guard
# ---------------------------------------------------------------------------

def test_every_assumption_has_a_named_range_for_every_forecast_year(apple_book,
                                                                   apple_spec):
    names = set(apple_book.defined_names)
    years = len(ts.forecast_periods(apple_spec))

    assert len(apple_spec.assumptions) * years + years == len(names)
    for assumption in apple_spec.assumptions:
        for index in range(years):
            assert excel.assumption_name(assumption.key, index) in names
    for index in range(years):
        assert excel.ready_name(index) in names


def test_a_defined_name_is_shaped_the_way_excel_requires(apple_book):
    """A malformed one is a repair dialog, which is what risk R4 is about."""
    for name in apple_book.defined_names:
        assert name[0].isalpha() or name[0] == "_"
        assert all(ch.isalnum() or ch == "_" for ch in name)
        assert not name[0].isdigit()


def test_every_assumption_cell_is_blank(apple_book):
    """The tool does the plumbing; the analyst does the thinking."""
    sheet = apple_book["Assumptions"]
    for name, defined in apple_book.defined_names.items():
        if "_ready_" in name:
            continue
        for title, coordinate in defined.destinations:
            assert apple_book[title][coordinate].value is None
            assert apple_book[title][coordinate].fill.fgColor.rgb.endswith(
                excel.FILL_INPUT)
    assert sheet["A1"].value.endswith("Every cell is yours to fill.")


def test_the_readiness_cell_requires_this_year_and_every_earlier_one(apple_book,
                                                                    apple_spec):
    """A chain, so a ready column can never reference a blank one beside it."""
    sheet = apple_book["Assumptions"]
    row = _find_row(sheet, "Year ready to model?")
    first = sheet.cell(row=row, column=excel.DEFAULT_LAYOUT.first_data_column)
    second = sheet.cell(row=row, column=excel.DEFAULT_LAYOUT.first_data_column + 1)

    assert first.value.startswith("=IF(OR(ISBLANK(asm_rev_growth_y1)")
    assert first.value.count("ISBLANK(") == len(apple_spec.assumptions)
    assert second.value.startswith("=IF(asm_ready_y1,")


def test_every_forecast_cell_is_guarded_by_its_years_readiness(apple_book):
    """The property the acceptance checklist tests by hand, asserted in the file."""
    guarded = 0
    for title in ("Income Statement", "Balance Sheet", "Cash Flow"):
        sheet = apple_book[title]
        for period_label in ("FY2026E", "FY2027E", "FY2028E"):
            year = int(period_label[2:6]) - 2025
            column = _column_for(sheet, period_label)
            for row in range(excel.DEFAULT_LAYOUT.first_data_row,
                             sheet.max_row + 1):
                value = sheet.cell(row=row, column=column).value
                if value is None:
                    continue
                assert value.startswith("=IF(asm_ready_y{},".format(year)), value
                assert value.endswith(',"")')
                guarded += 1
    assert guarded > 100


def test_a_row_with_no_forecast_leaves_its_forecast_cells_empty(apple_book):
    """An empty cell contributes zero to the sums above it, which is the truth
    for a line the filer's balance sheet does not carry."""
    sheet = apple_book["Balance Sheet"]
    for period_label in ("FY2026E", "FY2027E", "FY2028E"):
        assert _cell(sheet, "Goodwill", period_label).value is None
    assert _cell(sheet, "Total Assets", "FY2026E").value.startswith("=IF(")


# ---------------------------------------------------------------------------
# Cross-sheet linkage
# ---------------------------------------------------------------------------

def test_the_cash_flow_reads_net_income_off_the_income_statement(apple_book):
    """Referenced rather than repeated, so the two cannot disagree."""
    cash_flow = apple_book["Cash Flow"]
    income = apple_book["Income Statement"]
    cell = _cell(cash_flow, "Net income (from the income statement)", "FY2023")

    assert cell.value == "='Income Statement'!{}{}".format(
        cell.column_letter, _find_row(income, "Net Income"))


def test_the_balance_sheet_reads_its_forecast_cash_off_the_cash_flow(apple_book):
    """The linkage the whole exercise is for."""
    balance = apple_book["Balance Sheet"]
    cash_flow = apple_book["Cash Flow"]
    cell = _cell(balance, "Cash and Equivalents", "FY2026E")

    assert "'Cash Flow'!{}{}".format(cell.column_letter,
                                     _find_row(cash_flow, "Cash, end of period")
                                     ) in cell.value


def test_opening_cash_points_at_the_prior_columns_balance_sheet(apple_book):
    cash_flow = apple_book["Cash Flow"]
    balance = apple_book["Balance Sheet"]
    cell = _cell(cash_flow, "Cash, beginning of period", "FY2023")
    prior = _column_for(balance, "FY2022")

    assert cell.value == "='Balance Sheet'!{}{}".format(
        balance.cell(row=1, column=prior).column_letter,
        _find_row(balance, "Cash and Equivalents"))


def test_the_first_column_has_no_opening_cash_because_nothing_precedes_it(apple_book):
    assert _cell(apple_book["Cash Flow"], "Cash, beginning of period",
                 "FY2021").value is None


# ---------------------------------------------------------------------------
# Schedules and checks
# ---------------------------------------------------------------------------

def test_each_schedule_carries_the_residual_that_makes_it_foot(apple_book):
    """A roll-forward that does not foot is worse than none.

    Opening property, plant and equipment plus capex less depreciation is not
    the closing balance for any real filer, because impairments, disposals and
    currency move it too and none of those is a registry item. The residual
    line is named for what sits in it.
    """
    sheet = apple_book["Schedules"]
    for label in ("Impairments, disposals, acquisitions and currency (implied)",
                  "Net issuance or repayment in the year (implied)",
                  "Share retirements and other equity movements (implied)",
                  "Effect of exchange rates on cash and other (implied)"):
        row = _find_row(sheet, label)
        assert sheet.cell(row=row,
                          column=_column_for(sheet, "FY2023")).value.startswith("=")


def test_a_schedule_line_with_no_prior_column_is_left_empty(apple_book):
    sheet = apple_book["Schedules"]
    row = _find_row(sheet, "Opening long-term debt")

    assert sheet.cell(row=row, column=_column_for(sheet, "FY2021")).value is None
    assert sheet.cell(row=row, column=_column_for(sheet, "FY2022")).value is not None


def test_a_tie_check_is_coloured_and_a_residual_is_not(apple_book):
    """Colouring a row red every year for a difference that is ordinary would
    train a reader to ignore the sheet."""
    sheet = apple_book["Checks"]
    rules = sheet.conditional_formatting
    coloured_rows = set()
    for entry in rules:
        for cells in str(entry.sqref).split():
            coloured_rows.add(int("".join(ch for ch in cells.split(":")[0]
                                          if ch.isdigit())))

    assert _find_row(sheet, "Balance check (assets less liabilities and equity)"
                     ) in coloured_rows
    assert _find_row(sheet, "Cash tie (cash flow close less balance sheet cash)"
                     ) in coloured_rows
    assert _find_row(sheet, "Retained earnings roll-forward residual"
                     ) not in coloured_rows


def test_honeywells_balance_check_is_written_but_not_coloured(honeywell_book):
    """It is zero by construction for this filer, and the sheet says so."""
    sheet = honeywell_book["Checks"]
    row = _find_row(sheet, "Balance check (assets less liabilities and equity)")
    label = sheet.cell(row=row, column=excel.DEFAULT_LAYOUT.label_column)
    coloured = {int("".join(ch for ch in str(entry.sqref).split(":")[0]
                            if ch.isdigit())) for entry in
                sheet.conditional_formatting}

    assert row not in coloured
    assert ts.FLAG_CHECK_NOT_AVAILABLE in label.comment.text
    assert sheet.cell(row=row,
                      column=_column_for(sheet, "FY2025")).value.startswith("=")


def test_the_checks_sheet_lists_what_the_scaffold_flags_about_this_filer(apple_book):
    """One line per flagged row, not one per flagged cell."""
    sheet = apple_book["Checks"]
    row = _find_row(sheet, "What this scaffold flags about this filer")
    lines = [sheet.cell(row=r, column=1).value
             for r in range(row + 1, sheet.max_row + 1)]

    assert any(ts.FLAG_PLUG_TOO_LARGE in line for line in lines)
    assert sum(1 for line in lines
               if "Other investing activities (plug to reported total)" in line) == 1


def test_the_checks_sheet_reports_coverage_for_every_balance_sheet_section(
        apple_book, apple_spec):
    """The measurement that replaced the warning, as live arithmetic.

    Written as the subtotal less the plug over the subtotal rather than as a
    percentage this code computed and typed in, so a reader who distrusts it can
    follow it to the two cells the balance sheet already shows.
    """
    sheet = apple_book["Checks"]
    heading = _find_row(sheet,
                        "How much of each balance sheet section this scaffold reaches")
    assert ts.COVERAGE_NOTE in sheet.cell(
        row=heading, column=excel.DEFAULT_LAYOUT.label_column).comment.text

    for offset, entry in enumerate(apple_spec.coverage):
        row = heading + 1 + offset
        assert sheet.cell(row=row, column=1).value == entry["total_row"]
        cell = sheet.cell(row=row, column=_column_for(sheet, "FY2025"))
        assert cell.value.startswith("=(")
        assert cell.value.count("!") == 3      # three references, all off-sheet
        assert cell.number_format == excel.FORMAT_PERCENT


# ---------------------------------------------------------------------------
# Presentation
#
# Everything here is about how a cell is displayed and nothing here is about
# what it holds. The rule the session that wrote them worked to is that the
# workbook could be regenerated and evaluate to the same numbers cell for cell,
# so a test in this block that could only pass by moving a value would be a test
# of the wrong thing.
# ---------------------------------------------------------------------------

def test_every_dollar_and_per_share_format_reserves_the_bracket_width(apple_book):
    """The alignment underscore, without which digits jump a character on sign.

    A negative prints inside brackets and a positive does not, so a positive
    needs the width of the closing bracket reserved after it or a column of
    numbers has its units column in two places.
    """
    for number_format in (excel.FORMAT_DOLLAR, excel.FORMAT_DOLLAR_MILLIONS,
                          excel.FORMAT_EPS, excel.FORMAT_SHARES_MILLIONS):
        positive, negative = number_format.split(";")
        assert positive.endswith("_)"), number_format
        assert negative.startswith("(") and negative.endswith(")"), number_format

    seen = set()
    for title in ("Income Statement", "Balance Sheet", "Cash Flow", "Schedules"):
        sheet = apple_book[title]
        for row in range(excel.DEFAULT_LAYOUT.first_data_row, sheet.max_row + 1):
            for column in range(excel.DEFAULT_LAYOUT.first_data_column,
                                sheet.max_column + 1):
                seen.add(sheet.cell(row=row, column=column).number_format)
    assert seen <= {"General", excel.FORMAT_DOLLAR_MILLIONS, excel.FORMAT_EPS,
                    excel.FORMAT_SHARES_MILLIONS}


def test_dollars_and_shares_display_in_millions_and_per_share_amounts_do_not(
        apple_book):
    """Scaled in the format, so the cell still holds what the filer tagged."""
    income = apple_book["Income Statement"]
    revenue = _cell(income, "Revenue", "FY2023")

    assert revenue.number_format == excel.FORMAT_DOLLAR_MILLIONS
    assert revenue.number_format.count(",,") == 2      # once per sign section
    assert revenue.value == 383_285 * MILLION          # and the value did not move

    assert _cell(income, "Shares Outstanding (Diluted)",
                 "FY2023").number_format == excel.FORMAT_SHARES_MILLIONS
    assert _cell(income, "EPS Diluted", "FY2023").number_format == excel.FORMAT_EPS
    assert ",," not in excel.FORMAT_EPS


def test_a_check_row_keeps_whole_dollars_so_a_small_residual_stays_visible(
        apple_book):
    """The exception the millions decision names, and the reason for it.

    A residual of a few thousand dollars displayed in millions rounds to a
    displayed zero, which would make a broken check look like a passing one.
    """
    sheet = apple_book["Checks"]
    row = _find_row(sheet, "Balance check (assets less liabilities and equity)")
    cell = sheet.cell(row=row, column=_column_for(sheet, "FY2025"))

    assert cell.number_format == excel.FORMAT_DOLLAR
    assert ",," not in cell.number_format


def test_each_sheet_says_what_scale_it_is_in(apple_book):
    """In the header region, where a reader of any model looks for it."""
    row = excel.DEFAULT_LAYOUT.title_row + 1
    column = excel.DEFAULT_LAYOUT.label_column

    for title in ("Income Statement", "Balance Sheet", "Cash Flow"):
        note = apple_book[title].cell(row=row, column=column).value
        assert "$ in millions" in note
        assert "Share counts in millions" in note
    assert "$ in millions" in apple_book["Schedules"].cell(row=row,
                                                           column=column).value
    assert "whole dollars" in apple_book["Checks"].cell(row=row, column=column).value
    assert "500 means 500 million" in apple_book["Assumptions"].cell(
        row=row, column=column).value


def test_a_formula_that_reads_another_sheet_is_green(apple_book):
    """The convention's one colour about where a number comes from.

    Blue still means the filer reported it and black still means this sheet
    computed it; green means the cell went to another sheet for it, which is
    the thing a reader most needs to see before they trust a number.
    """
    cash_flow = apple_book["Cash Flow"]
    linked = _cell(cash_flow, "Net income (from the income statement)", "FY2023")
    same_sheet = _cell(cash_flow, "Cash, end of period", "FY2023")
    reported = _cell(apple_book["Income Statement"], "Revenue", "FY2023")
    missing = _cell(apple_book["Balance Sheet"], "Goodwill", "FY2023")

    assert "!" in linked.value
    assert linked.font.color.rgb.endswith(excel.COLOR_LINK)
    assert "!" not in same_sheet.value
    assert same_sheet.font.color.rgb.endswith(excel.COLOR_DERIVED)
    assert reported.font.color.rgb.endswith(excel.COLOR_REPORTED)
    assert missing.font.color.rgb.endswith(excel.COLOR_MISSING)
    assert missing.font.italic


def test_the_colour_is_decided_by_the_formula_and_not_by_a_list_of_rows(apple_book):
    """Every cell that reaches off its sheet, not the six the plan happened to name.

    A list would go stale the first time a row was added and the colour would
    then be wrong about the one cell a reader most needs it to be right about.
    """
    for title in ("Income Statement", "Balance Sheet", "Cash Flow", "Schedules",
                  "Checks"):
        sheet = apple_book[title]
        for row in sheet.iter_rows(min_row=excel.DEFAULT_LAYOUT.first_data_row):
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                expected = (excel.COLOR_LINK if "!" in cell.value
                            else excel.COLOR_DERIVED)
                assert cell.font.color.rgb.endswith(expected), (
                    title, cell.coordinate, cell.value)


def test_a_detail_row_is_indented_and_its_label_is_not_changed(apple_spec, apple_book):
    """Indented by openpyxl, not by spaces, so a downstream read is unchanged."""
    sheet = apple_book["Balance Sheet"]
    subtotal = sheet.cell(row=_find_row(sheet, "Total Current Assets"), column=1)
    detail = sheet.cell(row=_find_row(sheet, "Accounts Receivable"), column=1)

    assert subtotal.alignment.indent == 0
    assert detail.alignment.indent == 1

    for code, sheet_name, _block in ts.STATEMENT_BLOCKS:
        written = apple_book[sheet_name]
        for row in ts.rows_for(apple_spec, code):
            label = written.cell(row=_find_row(written, row.label), column=1).value
            assert label == row.label
            assert label == label.strip()


def test_a_subtotal_is_ruled_off_and_each_statement_is_closed_with_a_double(
        apple_book):
    """Single top on a subtotal, double bottom on the row the statement ends on."""
    expected_bottom = {"Income Statement": "Net Income",
                       "Balance Sheet": "Total Equity",
                       "Cash Flow": "Cash, end of period"}
    for title, bottom_label in expected_bottom.items():
        sheet = apple_book[title]
        row = _find_row(sheet, bottom_label)
        for column in (1, _column_for(sheet, "FY2025")):
            border = sheet.cell(row=row, column=column).border
            assert border.top.style == "thin", (title, column)
            assert border.bottom.style == "double", (title, column)

    sheet = apple_book["Balance Sheet"]
    ruled = sheet.cell(row=_find_row(sheet, "Total Assets"), column=1).border
    assert ruled.top.style == "thin"
    assert ruled.bottom is None or ruled.bottom.style is None


def test_the_memo_rows_after_a_bottom_line_are_not_ruled_off(apple_book):
    """Earnings per share sits below net income and is not part of the total.

    The double rule marks where a statement ends, so it has to fall on the
    bottom line rather than on the last row that happens to be written.
    """
    income = apple_book["Income Statement"]
    for label in ("EBITDA", "EPS Diluted", "Shares Outstanding (Diluted)"):
        border = income.cell(row=_find_row(income, label), column=1).border
        assert border.top is None or border.top.style is None, label
        assert border.bottom is None or border.bottom.style is None, label

    balance = apple_book["Balance Sheet"]
    debt = balance.cell(row=_find_row(balance, "Total Debt"), column=1).border
    assert debt.bottom is None or debt.bottom.style is None


# ---------------------------------------------------------------------------
# The currency assumption, typed in millions
# ---------------------------------------------------------------------------

def test_a_currency_assumption_says_it_is_in_millions_and_is_not_scaled_again(
        apple_book):
    """The input holds millions, so its own format must not divide by a million."""
    sheet = apple_book["Assumptions"]
    row = _find_row(sheet, "Net long-term debt issuance ($ in millions)")
    cell = sheet.cell(row=row, column=excel.DEFAULT_LAYOUT.first_data_column)

    assert cell.value is None
    assert cell.number_format == excel.FORMAT_DOLLAR
    assert ",," not in cell.number_format


def test_a_percent_or_days_assumption_is_labelled_and_formatted_as_it_was(apple_book):
    """Only currency has a scale to get wrong, so only currency changed."""
    sheet = apple_book["Assumptions"]
    for label, number_format in (("Revenue growth", excel.FORMAT_PERCENT),
                                 ("Days sales outstanding", excel.FORMAT_DAYS)):
        row = _find_row(sheet, label)
        assert sheet.cell(row=row, column=excel.DEFAULT_LAYOUT.first_data_column
                          ).number_format == number_format


def test_every_formula_that_consumes_a_currency_assumption_multiplies_it_back(
        apple_book, apple_spec):
    """The arithmetic downstream stays in raw dollars, like the rest of the file."""
    currency = [a for a in apple_spec.assumptions
                if a.unit in excel.ASSUMPTION_SCALE]
    assert currency, "no currency assumption to check"

    found = 0
    for title in ("Income Statement", "Balance Sheet", "Cash Flow", "Schedules"):
        sheet = apple_book[title]
        for row in sheet.iter_rows(min_row=excel.DEFAULT_LAYOUT.first_data_row):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for assumption in currency:
                    for index in range(len(ts.forecast_periods(apple_spec))):
                        name = excel.assumption_name(assumption.key, index)
                        if name not in cell.value:
                            continue
                        assert "({}*1000000)".format(name) in cell.value, cell.value
                        found += 1
    assert found >= 2      # the debt balance and the financing line, at least


def test_a_percent_assumption_is_never_multiplied(apple_book):
    """A scale applied where there is none is the same bug in the other direction."""
    sheet = apple_book["Income Statement"]
    cell = _cell(sheet, "Revenue", "FY2026E")

    assert "asm_rev_growth_y1" in cell.value
    assert "asm_rev_growth_y1*" not in cell.value


# ---------------------------------------------------------------------------
# Source tags
# ---------------------------------------------------------------------------

def test_the_source_tags_sheet_holds_one_line_per_historical_value(apple_book,
                                                                  apple_spec):
    sheet = apple_book["Source Tags"]
    expected = len(apple_spec.rows) * len(ts.historical_periods(apple_spec))

    assert sheet.max_row == expected + 1
    assert [sheet.cell(row=1, column=c).value for c in range(1, 9)] == [
        "Statement", "Line item", "Period", "State", "Tag", "Filed", "Accession",
        "Note"]


def test_every_reported_line_of_the_source_sheet_names_a_tag(apple_book):
    sheet = apple_book["Source Tags"]
    reported = 0
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=4).value != ts.CELL_REPORTED:
            continue
        assert sheet.cell(row=row, column=5).value
        assert sheet.cell(row=row, column=7).value
        reported += 1
    assert reported > 100


# ---------------------------------------------------------------------------
# Nothing about the shape is hardcoded
# ---------------------------------------------------------------------------

def test_the_sheet_names_come_from_the_layout_and_the_formulas_follow(apple_spec,
                                                                     tmp_path):
    """Rename every sheet and the cross-sheet references rename with them.

    A writer that hardcodes "Income Statement" in a formula string passes every
    other test in this module and breaks the moment a caller wants a different
    arrangement, which Phases 3 and 4 will.
    """
    sheets = excel.SheetNames(assumptions="Inputs", income_statement="P&L",
                              balance_sheet="BS", cash_flow="CF",
                              schedules="Sched", checks="Ties", source_tags="Sources")
    layout = excel.DEFAULT_LAYOUT._replace(sheets=sheets)
    workbook, _path = _reread(apple_spec, tmp_path, "renamed.xlsx", layout)

    assert workbook.sheetnames == ["Inputs", "P&L", "BS", "CF", "Sched", "Ties",
                                   "Sources"]
    cell = _cell(workbook["CF"], "Net income (from the income statement)", "FY2023")
    assert cell.value.startswith("='P&L'!")
    assert "Income Statement" not in cell.value


def test_the_first_data_column_comes_from_the_layout_too(apple_spec, tmp_path):
    layout = excel.DEFAULT_LAYOUT._replace(first_data_column=4, label_column=2)
    workbook, _path = _reread(apple_spec, tmp_path, "shifted.xlsx", layout)
    sheet = workbook["Income Statement"]

    assert sheet.cell(row=layout.header_row, column=4).value == "FY2021"
    assert sheet.cell(row=layout.first_data_row, column=2).value == "Revenue"
    assert sheet.cell(row=layout.first_data_row, column=4).value == 365_817 * MILLION


def test_the_column_count_comes_from_the_spec(apple_spec, tmp_path):
    spec = _spec(APPLE, history_years=3, forecast_years=1)
    workbook, _path = _reread(spec, tmp_path, "short.xlsx")
    sheet = workbook["Income Statement"]
    layout = excel.DEFAULT_LAYOUT

    labels = [sheet.cell(row=layout.header_row, column=c).value
              for c in range(layout.first_data_column, layout.first_data_column + 5)]
    assert labels == ["FY2023", "FY2024", "FY2025", "FY2026E", None]
    assert len(workbook.defined_names) == len(spec.assumptions) + 1
