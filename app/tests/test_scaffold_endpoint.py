"""test_scaffold_endpoint.py -- POST /api/scaffold/three-statement, and the button.

V2_PLAN 2.3 is one endpoint and one button, and the endpoint's job is to
compose what Sessions 5 built rather than to decide anything itself. So the
first thing tested here is exactly that: a workbook the endpoint writes is cell
for cell the workbook the library writes, which is what stops scaffold rules
leaking into app.py the next time something needs adjusting.

The second thing is the refusals. A bank, an insurer and an IFRS-only filer get
the scope gate's own sentence and no file at all. Not a workbook with a warning
in it, and not an empty one: the whole response is the refusal, and there is
nothing on disk afterwards for anything else to pick up by mistake.

Every test writes into a temporary exports directory and none of them touches
the network.
"""

import json
import os
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

import app as flask_app
import edgar_api
import line_items
import xbrl_extractor as xbrl
from scaffold import excel as scaffold_excel
from scaffold import three_statement as ts

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")

APPLE = (320193, "3571")
HONEYWELL = (773840, "3728")
KROGER = (56873, "5411")
JPMORGAN = (19617, "6021")
SAP = (1000184, "7372")


def _facts(cik):
    with open(os.path.join(FIXTURES, "cik{}.json".format(cik)), encoding="utf-8") as h:
        return json.load(h)


@pytest.fixture
def offline(monkeypatch, tmp_path):
    """The app, wired to the committed fixtures and a throwaway exports folder.

    Returns a function that serves one company: everything the endpoint reaches
    for comes from disk, so a test that accidentally asks for a filer it did not
    set up fails rather than quietly going to EDGAR.
    """
    exports = tmp_path / "exports"
    exports.mkdir()
    monkeypatch.setattr(flask_app, "EXPORTS_DIR", str(exports))

    def serve(company):
        cik, sic = company
        monkeypatch.setattr(xbrl, "fetch_company_facts",
                            lambda requested: _facts(cik))
        monkeypatch.setattr(edgar_api, "get_company_meta",
                            lambda requested: {"sic": sic})
        return flask_app.app.test_client()

    serve.exports = exports
    return serve


def _post(client, **body):
    body.setdefault("format", "xlsx")
    return client.post("/api/scaffold/three-statement", json=body)


def _written(exports):
    return [os.path.join(root, name)
            for root, _dirs, files in os.walk(str(exports)) for name in files]


# ---------------------------------------------------------------------------
# The endpoint composes and decides nothing
# ---------------------------------------------------------------------------

def test_the_endpoint_writes_the_workbook_the_library_writes(offline, tmp_path):
    """Cell for cell, against the same spec built directly.

    If this ever fails, something about what a scaffold contains has been
    decided in app.py, which is the one thing the split between the model spec
    and the writer exists to prevent.
    """
    client = offline(APPLE)
    response = _post(client, cik="320193", years=5)
    assert response.status_code == 200, response.get_json()

    from_endpoint = openpyxl.load_workbook(
        os.path.join(str(offline.exports), *response.get_json()["download_url"]
                     .split("/exports/")[1].split("/")))

    direct = str(tmp_path / "direct.xlsx")
    scaffold_excel.write_workbook(
        ts.build_model(320193, _facts(320193), "3571", history_years=5,
                       forecast_years=3), direct)
    expected = openpyxl.load_workbook(direct)

    assert from_endpoint.sheetnames == expected.sheetnames
    compared = 0
    for name in expected.sheetnames:
        mine, theirs = from_endpoint[name], expected[name]
        assert (mine.max_row, mine.max_column) == (theirs.max_row, theirs.max_column)
        for row in range(1, theirs.max_row + 1):
            for column in range(1, theirs.max_column + 1):
                assert mine.cell(row=row, column=column).value == \
                    theirs.cell(row=row, column=column).value, \
                    "{}!{}{}".format(name, column, row)
                compared += 1
    assert compared > 1000
    assert set(from_endpoint.defined_names) == set(expected.defined_names)


def test_the_workbook_has_the_seven_sheets_the_plan_names(offline):
    client = offline(HONEYWELL)
    response = _post(client, cik="773840")
    assert response.status_code == 200

    path = _written(offline.exports)[0]
    assert openpyxl.load_workbook(path).sheetnames == [
        "Assumptions", "Income Statement", "Balance Sheet", "Cash Flow",
        "Schedules", "Checks", "Source Tags"]


def test_the_response_says_what_the_workbook_holds(offline):
    """Columns and flags, so the Checks sheet is not the first place a reader
    learns there was something to read."""
    client = offline(KROGER)
    payload = _post(client, cik="56873", years=5).get_json()

    assert payload["status"] == "ok"
    assert payload["entity"] == "KROGER CO"
    assert payload["historical"] == ["FY2021", "FY2022", "FY2023", "FY2024",
                                     "FY2025"]
    assert payload["forecast"] == ["FY2026E", "FY2027E", "FY2028E"]
    assert payload["filename"].endswith(".xlsx")
    assert payload["download_url"].startswith("/exports/")
    assert payload["scope"]["in_scope"] is True

    kinds = {flag["flag_type"] for flag in payload["flags"]}
    # Kroger tags no inventory and no SG&A, so the rows that would forecast
    # them say so rather than being silently absent.
    assert ts.FLAG_NO_REPORTED_HISTORY in kinds
    assert all(flag["message"] for flag in payload["flags"])


def test_the_years_asked_for_are_the_years_written(offline):
    client = offline(APPLE)

    for years in (1, 3, 5):
        payload = _post(client, cik="320193", years=years).get_json()
        assert len(payload["historical"]) == years
        assert payload["historical"][-1] == "FY2025"


def test_no_forecast_columns_is_a_legitimate_request(offline):
    """A historical workbook with nothing modelled is a thing to want, and it
    must not be the same code path as a forecast that failed."""
    client = offline(APPLE)
    payload = _post(client, cik="320193", years=3, forecast_years=0).get_json()

    assert payload["forecast"] == []
    sheet = openpyxl.load_workbook(_written(offline.exports)[0])["Assumptions"]
    assert sheet.cell(row=scaffold_excel.DEFAULT_LAYOUT.header_row,
                      column=scaffold_excel.DEFAULT_LAYOUT.first_data_column
                      ).value is None


def test_the_written_file_is_served_back(offline):
    client = offline(APPLE)
    payload = _post(client, cik="320193", years=2).get_json()

    served = client.get(payload["download_url"])
    assert served.status_code == 200
    assert len(served.data) == os.path.getsize(_written(offline.exports)[0])


# ---------------------------------------------------------------------------
# Refusals
# ---------------------------------------------------------------------------

def test_a_bank_is_refused_with_the_scope_gates_own_message(offline):
    """JPMorgan, SIC 6021. The exact sentence V2_PLAN 1.4 fixes, not a paraphrase."""
    client = offline(JPMORGAN)
    response = _post(client, cik="19617")
    payload = response.get_json()

    assert response.status_code == 422
    assert payload["error"] == ("Bank and insurance financial statements do not "
                                "fit the standard three-statement template; "
                                "Edgardly will not generate a scaffold for this "
                                "company. SIC code 6021 is in the bank range.")
    assert payload["scope"]["reason"] == line_items.SCOPE_FINANCIAL_SIC
    assert payload["entity"]


def test_an_ifrs_filer_is_refused_with_its_own_message(offline):
    """SAP reports no us-gaap facts at all, which is a different refusal."""
    client = offline(SAP)
    response = _post(client, cik="1000184")
    payload = response.get_json()

    assert response.status_code == 422
    assert payload["error"].startswith("This company reports under IFRS, not US GAAP")
    assert "limit of this tool, not a gap in the company's filings" in payload["error"]
    assert payload["scope"]["reason"] == line_items.SCOPE_IFRS_ONLY


def test_an_insurer_is_refused_on_its_sic_alone(offline, monkeypatch):
    """No insurer fixture exists, so Apple's payload is served under an insurer's
    SIC. That is the honest test of a deterministic gate: the refusal is the
    code, and a filer with perfectly ordinary tagging is still refused."""
    monkeypatch.setattr(xbrl, "fetch_company_facts", lambda cik: _facts(320193))
    monkeypatch.setattr(edgar_api, "get_company_meta", lambda cik: {"sic": "6311"})

    response = _post(flask_app.app.test_client(), cik="320193")

    assert response.status_code == 422
    assert "SIC code 6311 is in the insurance range" in response.get_json()["error"]


def test_a_refusal_writes_no_file(offline):
    """Not an empty workbook and not one with the refusal typed into it."""
    client = offline(JPMORGAN)
    assert _post(client, cik="19617").status_code == 422

    assert _written(offline.exports) == []


def test_a_failed_sic_lookup_does_not_stop_a_scaffold(offline, monkeypatch):
    """The gate's shape heuristic still runs, and Apple is still Apple.

    The submissions API is a second request and it is allowed to fail. Refusing
    to build anything when it does would make an unrelated outage look like a
    verdict about the company.
    """
    def boom(cik):
        raise RuntimeError("submissions API unavailable")

    client = offline(APPLE)
    monkeypatch.setattr(edgar_api, "get_company_meta", boom)
    payload = _post(client, cik="320193", years=2).get_json()

    assert payload["status"] == "ok"
    assert payload["scope"]["detail"]["sic_lookup"] == "unavailable"
    assert payload["scope"]["detail"]["sic"] is None


# ---------------------------------------------------------------------------
# Bad requests
# ---------------------------------------------------------------------------

def test_a_missing_cik_is_a_bad_request(offline):
    client = offline(APPLE)
    response = client.post("/api/scaffold/three-statement", json={})

    assert response.status_code == 400
    assert response.get_json()["error"] == "cik is required"


@pytest.mark.parametrize("body,fragment", [
    ({"years": "five"}, "must be integers"),
    ({"years": 0}, "years must be between 1 and 20"),
    ({"years": 21}, "years must be between 1 and 20"),
    ({"forecast_years": -1}, "forecast_years must be between 0 and 20"),
])
def test_a_nonsense_year_count_is_a_bad_request(offline, body, fragment):
    client = offline(APPLE)
    response = _post(client, cik="320193", **body)

    assert response.status_code == 400
    assert fragment in response.get_json()["error"]


def test_csv_is_refused_because_a_scaffold_is_not_a_table(offline):
    """Flattening seven linked sheets to values would destroy the only thing
    the workbook is for, so it is refused rather than quietly written."""
    client = offline(APPLE)
    response = client.post("/api/scaffold/three-statement",
                           json={"cik": "320193", "format": "csv"})

    assert response.status_code == 400
    assert "cannot be written as CSV" in response.get_json()["error"]
    assert _written(offline.exports) == []


def test_a_failure_reaching_edgar_is_reported_rather_than_swallowed(monkeypatch):
    def boom(cik):
        raise RuntimeError("EDGAR returned 503")

    monkeypatch.setattr(xbrl, "fetch_company_facts", boom)
    response = _post(flask_app.app.test_client(), cik="320193")

    assert response.status_code == 500
    assert "503" in response.get_json()["error"]


# ---------------------------------------------------------------------------
# The button
# ---------------------------------------------------------------------------

def _template():
    path = os.path.join(os.path.dirname(__file__), "..", "templates", "index.html")
    with open(path, encoding="utf-8") as handle:
        return handle.read()


def test_the_xbrl_view_carries_a_button_that_calls_the_endpoint():
    source = _template()

    assert 'id="xbrl-scaffold-btn"' in source
    assert "Build 3-Statement Model" in source
    assert "'/api/scaffold/three-statement'" in source


def test_the_button_surfaces_a_refusal_instead_of_an_error():
    """A 422 is the gate speaking, not a failure, and the page must not dress it
    up as one. The refusal element is separate from the status line for that
    reason."""
    source = _template()

    assert "status === 422" in source
    assert 'id="xbrl-scaffold-refusal"' in source
    assert "xbrlShowScaffoldRefusal" in source
