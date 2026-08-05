import os
import subprocess
import sys
import webbrowser
import threading
import datetime
import json
import edgar_api
import line_items
import periods
import xbrl_extractor as xbrl
import peer_comparison as pc
from scaffold import excel as scaffold_excel
from scaffold import three_statement as scaffold_model
from flask import Flask, render_template, request, jsonify, send_from_directory, Response, stream_with_context

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DOWNLOADS_DIR = os.path.join(BASE_DIR, "downloads")
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DOWNLOADS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)


def _xbrl_format_value(value, unit):
    if value is None:
        return None
    if unit == "USD":
        a = abs(value)
        if a >= 1e9:
            return "${:.2f}B".format(value / 1e9)
        if a >= 1e6:
            return "${:.1f}M".format(value / 1e6)
        return "${:,.0f}".format(value)
    if unit in ("USD/shares",):
        return "${:.2f}".format(value)
    if unit == "shares":
        a = abs(value)
        if a >= 1e9:
            return "{:.2f}B".format(value / 1e9)
        if a >= 1e6:
            return "{:.1f}M".format(value / 1e6)
        return "{:,.0f}".format(value)
    return "{:g}".format(value)


# ---------------------------------------------------------------------------
# XBRL export: line-item classification and Excel format constants
# ---------------------------------------------------------------------------

# Canonical definitions live in line_items.py, shared with peer_comparison.
# Re-exported under the private names the rest of this module already uses.
_DOLLAR_LINE_ITEMS = line_items.DOLLAR_LINE_ITEMS
_EPS_LINE_ITEMS = line_items.EPS_LINE_ITEMS
_SHARE_LINE_ITEMS = line_items.SHARE_LINE_ITEMS

# Accounting-style format: positives with trailing space (aligns with closing paren on negatives),
# negatives in parentheses, zero as dash.
_XLSX_FMT_DOLLAR = '#,##0_);(#,##0);"-"'
_XLSX_FMT_EPS    = '#,##0.00_);(#,##0.00);"-"'
_XLSX_FMT_SHARES = '#,##0_);(#,##0);"-"'

# Rows that receive a border in Excel exports (financial modeling convention)
_SINGLE_BORDER_ROW_ITEMS = frozenset({"Gross Profit", "Operating Income"})
_DOUBLE_BORDER_ROW_ITEMS = frozenset({"Net Income"})

# ---------------------------------------------------------------------------
# Shared Excel sanity-check helper
# ---------------------------------------------------------------------------

def _write_xlsx_sanity_checks(ws, start_row, row_map, data_col_idxs, dollar_label, BF):
    """Append formula-driven sanity-check rows to a worksheet.

    row_map:       {line_item_name: excel_row_number}  -- built while writing data rows
    data_col_idxs: ordered list of 1-based column indices that hold financial data

    Each check cell contains a live Excel formula so that editing any cell in the
    sheet immediately recalculates the result.  Conditional-formatting rules colour
    the result green (within 5% tolerance), red (fails), or grey (any input is N/A
    or non-numeric).

    Balance-Sheet:  Total Assets − Total Liabilities − Total Equity  (should be 0)
    Gross-Profit:   Revenue − Cost of Revenue − Gross Profit          (should be 0)

    Returns the row number immediately after the last written row.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    TOLERANCE = 0.05
    CHECKS = [
        # (label, [(line_item, sign), ...], reference_item_for_tolerance)
        ("BS Check: Assets − Liabilities − Equity",
         [("Total Assets", 1), ("Total Liabilities", -1), ("Total Equity", -1)],
         "Total Assets"),
        ("GP Check: Revenue − COGS − Gross Profit",
         [("Revenue", 1), ("Cost of Revenue", -1), ("Gross Profit", -1)],
         "Revenue"),
    ]

    label_font   = Font(name=BF, size=11, italic=True)
    missing_font = Font(name=BF, color="AAAAAA", italic=True, size=11)

    _green_fill = PatternFill("solid", fgColor="D4EDDA")
    _green_font = Font(color="155724", name=BF, size=11)
    _red_fill   = PatternFill("solid", fgColor="F8D7DA")
    _red_font   = Font(color="721C24", name=BF, size=11)
    _grey_fill  = PatternFill("solid", fgColor="F5F5F5")
    _grey_font  = Font(color="AAAAAA", name=BF, size=11, italic=True)

    hdr = ws.cell(start_row, 1, "Sanity Checks")
    hdr.font = Font(name=BF, bold=True, size=11, color="003366")

    cur = start_row + 1
    for check_name, items_signs, ref_item in CHECKS:
        ws.cell(cur, 1, check_name).font = label_font

        # Resolve row numbers; if any required item is absent from the sheet
        # entirely, fall back to a static label.
        item_rows = {li: row_map.get(li) for li, _ in items_signs}
        ref_row   = row_map.get(ref_item)

        if any(r is None for r in item_rows.values()):
            for col_idx in data_col_idxs:
                c = ws.cell(cur, col_idx)
                c.value     = "Item not exported"
                c.font      = missing_font
                c.alignment = Alignment(horizontal="right")
        else:
            for col_idx in data_col_idxs:
                col = get_column_letter(col_idx)
                c   = ws.cell(cur, col_idx)

                # ISNUMBER guards prevent #VALUE! when a cell contains "N/A"/"Not reported"
                guards = ",".join(
                    f"ISNUMBER({col}{item_rows[li]})" for li, _ in items_signs
                )

                # Build arithmetic expression  e.g.  B7-B8-B9
                expr = "".join(
                    (f"{col}{item_rows[li]}" if i == 0
                     else f"+{col}{item_rows[li]}" if sign > 0
                     else f"-{col}{item_rows[li]}")
                    for i, (li, sign) in enumerate(items_signs)
                )

                c.value         = f'=IF(AND({guards}),{expr},"N/A")'
                c.number_format = _XLSX_FMT_DOLLAR
                c.alignment     = Alignment(horizontal="right")

            # Conditional formatting applied to the whole check row
            if data_col_idxs and ref_row is not None:
                first = get_column_letter(data_col_idxs[0])
                last  = get_column_letter(data_col_idxs[-1])
                rng   = f"{first}{cur}:{last}{cur}"
                # Anchor cell for CF formulas (top-left of range; column shifts automatically)
                a   = f"{first}{cur}"         # check result
                ref = f"{first}{ref_row}"     # reference component (e.g. Total Assets)

                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f"AND(ISNUMBER({a}),ABS({a})<={TOLERANCE}*ABS({ref}))"],
                    fill=_green_fill, font=_green_font,
                ))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f"AND(ISNUMBER({a}),ABS({a})>{TOLERANCE}*ABS({ref}))"],
                    fill=_red_fill, font=_red_font,
                ))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f"NOT(ISNUMBER({a}))"],
                    fill=_grey_fill, font=_grey_font,
                ))

        cur += 1

    ws.cell(cur, 1,
            f"Sanity checks ({dollar_label}):  "
            "green ✓ = within 5% tolerance   "
            "red ✗ = check fails   "
            "grey = data unavailable for this period"
    ).font = Font(name=BF, color="666666", italic=True, size=10)
    return cur + 1


def _detect_dollar_scale(rows, columns):
    """Return (factor, label) for dollar scaling based on most recent Revenue.

    Falls back to Total Assets if Revenue has no data in the displayed column range
    (this can happen when a company switched XBRL revenue tags mid-history and the
    older tag only covers pre-filter years).
    """
    for candidate in ("Revenue", "Total Assets"):
        for row in rows:
            if row["line_item"] == candidate:
                for col in reversed(columns):
                    cell = row["cells"].get(col["key"])
                    if cell and cell.get("value") is not None:
                        return line_items.dollar_scale_for(cell["value"])
    return line_items.DEFAULT_DOLLAR_SCALE


def _detect_share_scale(rows, columns):
    """Return (factor, label) for share-count scaling based on most recent share count."""
    for name in ("Shares Outstanding (Diluted)", "Shares Outstanding (Basic)"):
        for row in rows:
            if row["line_item"] == name:
                for col in reversed(columns):
                    cell = row["cells"].get(col["key"])
                    if cell and cell.get("value") is not None:
                        return line_items.share_scale_for(cell["value"])
    return line_items.DEFAULT_SHARE_SCALE


# Items the tables may fill by arithmetic when the filer tags no value of its
# own, in the order they have to be computed. The list lives in the registry so
# the peer table cannot disagree with this one about what a number is.
_DERIVED_UI_ITEMS = line_items.DERIVED_UI_ITEMS


def _company_sic(cik):
    """The filer's SIC code, and whether the lookup that found it worked.

    It lives in the submissions API rather than in companyfacts, so it is a
    second request. It is cached and it is allowed to fail: a company should
    still load with its SIC unknown rather than not load at all, and the scope
    gate's shape heuristic still runs without it.
    """
    try:
        return (edgar_api.get_company_meta(cik) or {}).get("sic") or None, "submissions"
    except Exception:
        return None, "unavailable"


def _company_scope(cik, facts):
    """Run the scope gate for a company, JSON-ready."""
    sic, lookup = _company_sic(cik)
    verdict = line_items.is_in_scope(sic, facts)
    detail = dict(verdict.detail)
    detail["sic_lookup"] = lookup
    return {
        "in_scope": verdict.in_scope,
        "reason": verdict.reason,
        "message": verdict.message,
        "detail": detail,
    }


def _reported_cell(dp, end, fp):
    return {
        "value": dp["value"],
        "formatted": _xbrl_format_value(dp["value"], dp.get("unit", "")),
        "unit": dp.get("unit"),
        "start": dp.get("start"),
        "end": end,
        "fp": fp,
        "tag": dp.get("tag"),
        "filed": dp.get("filed"),
        "flags": [],
        "provenance": xbrl.reported_provenance(dp),
    }


def _derivation_input(name, cell):
    """One entry of a derived value's provenance, describing where it came from.

    A reported input names its tag and its filing. A derived input carries its
    own formula and its own inputs instead, so a total built on a subtotal can
    still be checked all the way down to values a filer tagged.
    """
    prov = cell.get("provenance") or {}
    entry = {
        "name": name,
        "value": cell.get("value"),
        "tag": cell.get("tag"),
        "filed": cell.get("filed"),
        "accession": prov.get("accession"),
    }
    if prov.get("state") == xbrl.PROVENANCE_DERIVED:
        entry["formula"] = prov.get("formula")
        entry["inputs"] = prov.get("inputs", [])
    return entry


def _fill_derived_cells(rows, columns):
    """Compute the values the filer left untagged but the table can prove.

    A derived value is arithmetic on values from the same period in the same
    unit, and nothing else. An input must itself be reported, or be a value
    this function has already derived from reported ones: Total Debt is the
    sum of a long-term balance a filer tagged and a short-term balance built
    from the current-liabilities lines it tagged, and there is no way to state
    that in one step for the filers who report the parts and not the total.
    Nesting stops there, at one level, and every leaf is reported. A period
    that fails any of those conditions stays missing.
    """
    by_item = {row["line_item"]: row for row in rows}
    usable = (xbrl.PROVENANCE_REPORTED, xbrl.PROVENANCE_DERIVED)

    for name in _DERIVED_UI_ITEMS:
        rule = line_items.DERIVATIONS.get(name)
        target = by_item.get(name)
        if rule is None or target is None:
            continue
        sources = [(input_name, by_item.get(input_name)) for input_name, _sign in rule.inputs]

        for col in columns:
            key = col["key"]
            existing = target["cells"].get(key)
            if existing is not None and existing.get("value") is not None:
                continue

            input_cells = []
            for input_name, row in sources:
                cell = (row or {}).get("cells", {}).get(key)
                prov = (cell or {}).get("provenance") or {}
                if (cell is None or cell.get("value") is None
                        or prov.get("state") not in usable):
                    continue
                input_cells.append((input_name, cell))

            values = {n: c["value"] for n, c in input_cells}
            used = line_items.inputs_used(name, values)
            if not used:
                continue
            input_cells = [(n, c) for n, c in input_cells if n in used]

            units = {cell.get("unit") for _name, cell in input_cells}
            if len(units) != 1:
                continue
            unit = units.pop()

            value = line_items.derive(name, values)
            if value is None:
                continue

            target["cells"][key] = {
                "value": value,
                "formatted": _xbrl_format_value(value, unit),
                "unit": unit,
                "start": input_cells[0][1].get("start"),
                "end": key,
                "fp": col.get("fp"),
                "tag": None,
                "filed": None,
                "flags": [],
                "provenance": xbrl.derived_provenance(
                    line_items.formula_for(name, values),
                    [_derivation_input(n, c) for n, c in input_cells]),
            }


def _missing_derivation_inputs(name, rows_by_item, key):
    """Which of a derived row's inputs had no value for one period."""
    rule = line_items.DERIVATIONS.get(name)
    if rule is None:
        return []
    absent = []
    for input_name, _sign in rule.inputs:
        cell = (rows_by_item.get(input_name) or {}).get("cells", {}).get(key)
        if cell is None or cell.get("value") is None:
            absent.append(input_name)
    return absent


def _fill_underivable_cells(rows, columns, cik, pointers):
    """Explain a hole in a row that is arithmetic rather than a tag.

    Runs before the derivation-input rows are dropped, because the explanation
    is which of them was empty. A row like Total Debt has no XBRL tag anywhere
    and never will, so the ordinary "not tagged in XBRL" message would send a
    reader hunting for a line that no balance sheet carries. What is actually
    absent is a component, and naming it is both true and useful: JPMorgan
    reports short-term borrowings and no long-term debt, and that sentence is
    the whole story of why the row is blank.
    """
    rows_by_item = {row["line_item"]: row for row in rows}
    derived_only = [name for name in line_items.UI_LINE_ITEMS
                    if name in line_items.DERIVATIONS and name not in line_items.REGISTRY]

    for name in derived_only:
        row = rows_by_item.get(name)
        if row is None:
            continue
        for col in columns:
            key = col["key"]
            if row["cells"].get(key) is not None:
                continue
            # The column's own name. Built once, where the filer's fiscal-year
            # convention is known, so a pointer cannot send a reader to the
            # FY2026 10-K of a company whose table says FY2025.
            label = col.get("label") or ""
            row["cells"][key] = {
                "value": None,
                "formatted": None,
                "unit": None,
                "start": None,
                "end": key,
                "fp": col.get("fp"),
                "tag": None,
                "filed": None,
                "flags": [],
                "provenance": xbrl.missing_provenance(
                    name, label, cik, pointers.get(key),
                    xbrl.FLAG_DERIVATION_UNAVAILABLE,
                    _missing_derivation_inputs(name, rows_by_item, key)),
            }


def _fill_missing_cells(rows, columns, cik, pointers, tagged_ends=None):
    """Give every remaining hole a pointer to where the number would be.

    After this runs, every (row, column) in the table is exactly one of the
    three provenance states. A hole is a fact about the filing, so it says
    which statement of which filing to open rather than rendering as nothing.

    tagged_ends maps a line item to the period end dates the filer tagged it
    for at all. A hole whose end date is in that set is not an untagged item:
    it is a value this table could not confirm covers the period, which is a
    different sentence and a true one.
    """
    tagged_ends = tagged_ends or {}
    for row in rows:
        tagged = tagged_ends.get(row["line_item"], ())
        for col in columns:
            key = col["key"]
            if row["cells"].get(key) is not None:
                continue
            label = col.get("label") or ""
            flag = (xbrl.FLAG_PERIOD_UNRESOLVED if key in tagged
                    else xbrl.FLAG_NOT_TAGGED)
            row["cells"][key] = {
                "value": None,
                "formatted": None,
                "unit": None,
                "start": None,
                "end": key,
                "fp": col.get("fp"),
                "tag": None,
                "filed": None,
                "flags": [],
                "provenance": xbrl.missing_provenance(
                    row["line_item"], label, cik, pointers.get(key), flag),
            }


def _tag_summary(row, columns):
    """Name every tag the displayed values in one row came from, oldest first.

    A stitched row can span tag eras, and one tag_used cannot describe it
    honestly. Where a row has only one tag this reads exactly as it always did.
    """
    tags = []
    formulas = []
    for col in columns:
        prov = (row["cells"].get(col["key"]) or {}).get("provenance") or {}
        state = prov.get("state")
        if state == xbrl.PROVENANCE_DERIVED:
            formula = prov.get("formula")
            if formula and formula not in formulas:
                formulas.append(formula)
            continue
        if state != xbrl.PROVENANCE_REPORTED:
            continue
        tag = prov.get("tag")
        if tag and tag not in tags:
            tags.append(tag)
    if tags:
        return " -> ".join(tags)
    # A row with no tag anywhere is not a row with no source. Total Debt has
    # none by definition, and a filer that never tags gross profit has none
    # either; in both cases the arithmetic is what a reader wants named.
    if formulas:
        return "derived: {}".format(" -> ".join(formulas))
    return row.get("tag_used") or ""


def _build_xbrl_result(cik, start_year, end_year, period_type):
    facts = xbrl.fetch_company_facts(cik)
    entity = facts.get("entityName", str(cik))
    scope = _company_scope(cik, facts)
    # The displayed reported items, plus the ones only a displayed derivation
    # needs. The extra four never become rows; they are dropped below, once
    # Total Debt has been built out of them.
    extracted = [name for name in line_items.UI_LINE_ITEMS if name in line_items.REGISTRY]
    extracted += [name for name in line_items.DERIVATION_INPUT_ITEMS
                  if name not in extracted]
    raw = xbrl.extract_all_line_items(facts, extracted)
    deduped = xbrl.deduplicate_all_line_items(raw)
    all_flags = xbrl.validate_financials(deduped)

    # Which periods this filer reports, decided by dates rather than by EDGAR's
    # fiscal_period label. The same engine the peer table runs on, which is the
    # point: the two views used to disagree about whether a year existed, and a
    # balance sheet a later 10-Q had relabeled fell out of this one (V2_PLAN R5).
    confirmed = periods.period_ends(deduped, xbrl.TAG_MAP, period_type)
    in_range = {end: fp for end, fp in confirmed.items()
                if start_year <= int(end[:4]) <= end_year}

    # What this filer calls its own fiscal years. Zero for a calendar-year
    # filer, so most companies are named exactly as they were; Kroger's late
    # January year end is where the two conventions part.
    fy_offset = xbrl.fiscal_year_offset(facts)

    # A quarter is named for the fiscal year it sits in rather than the calendar
    # year it ends in, so the quarterly view and the annual view agree about the
    # same date. That needs the confirmed year ends, which for the annual view
    # are the columns themselves and for the quarterly view are a separate ask.
    annual_ends = (sorted(confirmed) if period_type == "annual"
                   else sorted(periods.period_ends(deduped, xbrl.TAG_MAP, periods.ANNUAL)))

    columns = []
    for end in sorted(in_range):
        fp = in_range[end]
        label = xbrl.period_label(end, fp, period_type, fy_offset, annual_ends)
        yr = int(end[:4]) - (fy_offset if period_type == "annual" else 0)
        columns.append({"key": end, "label": label, "fp": fp, "fy": yr})

    # Every displayed item, then the derivation inputs, which are dropped again
    # below. A displayed item with no chain (Total Debt) starts with no cells
    # and is filled entirely by _fill_derived_cells.
    rows = []
    for line_item in list(line_items.UI_LINE_ITEMS) + [
            name for name in extracted if name not in line_items.UI_LINE_ITEMS]:
        info = deduped.get(line_item) or {"data": [], "tag_used": None}
        tag_used = info.get("tag_used")
        item_flags = all_flags.get(line_item, [])
        cells = {}
        for end, dp in periods.points_by_end(info["data"], in_range, period_type).items():
            period_flags = [
                {"type": f["flag_type"], "msg": f["message"]}
                for f in item_flags if f.get("period_end") == end
            ]
            # The column's label, not the data point's own: the column is the
            # period, and the point was accepted because it covers that period.
            cells[end] = _reported_cell(dp, end, in_range[end])
            cells[end]["flags"] = period_flags
        rows.append({"line_item": line_item, "tag_used": tag_used, "cells": cells})

    # Provenance: derive what can be proven from reported values, then give
    # every remaining hole a pointer. Order matters -- a derivable period is a
    # derived value, not a missing one.
    _fill_derived_cells(rows, columns)
    _fill_underivable_cells(rows, columns, cik, xbrl.filing_pointers(facts))
    rows = [row for row in rows if row["line_item"] in line_items.UI_LINE_ITEMS]

    tagged_ends = {
        name: {dp.get("end") for dp in info.get("data", []) if dp.get("value") is not None}
        for name, info in deduped.items()
    }
    _fill_missing_cells(rows, columns, cik, xbrl.filing_pointers(facts), tagged_ends)
    for row in rows:
        row["tag_summary"] = _tag_summary(row, columns)

    return entity, columns, rows, scope


def _build_chart_data(entity, columns, rows, dollar_factor, dollar_label):
    """
    Convert _build_xbrl_result output into chart-ready series.

    Revenue and Net Income are scaled by dollar_factor and always included
    when present, marked flagged=True when the underlying cell is flagged.

    Gross Margin % and Net Margin % are computed only when BOTH inputs are
    present and NEITHER is flagged (flagged inputs make derived ratios
    unreliable; those periods appear as null/gap in the chart).

    Each point:  {"period_label": str, "period_end": str, "value": float|None,
                  "flagged": bool}
    """
    row_by_li = {r["line_item"]: r for r in rows}

    def _cell(li, end):
        row = row_by_li.get(li)
        if not row:
            return None, False
        cell = row["cells"].get(end)
        if not cell:
            return None, False
        val = cell.get("value")
        flagged = bool(cell.get("flags"))
        return val, flagged

    series = {k: [] for k in ("revenue", "net_income", "gross_margin_pct", "net_margin_pct")}

    for col in columns:
        end   = col["key"]
        label = col["label"]

        rev_val,  rev_flagged  = _cell("Revenue",      end)
        ni_val,   ni_flagged   = _cell("Net Income",   end)
        gp_val,   gp_flagged   = _cell("Gross Profit", end)

        safe_factor = dollar_factor if dollar_factor else 1

        def _scale(v):
            if v is None:
                return None
            return round(v / safe_factor, 2) if safe_factor != 1 else round(float(v), 2)

        series["revenue"].append({
            "period_label": label, "period_end": end,
            "value": _scale(rev_val), "flagged": rev_flagged,
        })
        series["net_income"].append({
            "period_label": label, "period_end": end,
            "value": _scale(ni_val), "flagged": ni_flagged,
        })

        # Gross Margin %: only when both present and neither flagged
        if (rev_val is not None and rev_val != 0
                and gp_val is not None
                and not rev_flagged and not gp_flagged):
            gm_pct = round(100.0 * gp_val / rev_val, 2)
        else:
            gm_pct = None

        series["gross_margin_pct"].append({
            "period_label": label, "period_end": end,
            "value": gm_pct, "flagged": False,
        })

        # Net Margin %: only when both present and neither flagged
        if (rev_val is not None and rev_val != 0
                and ni_val is not None
                and not rev_flagged and not ni_flagged):
            nm_pct = round(100.0 * ni_val / rev_val, 2)
        else:
            nm_pct = None

        series["net_margin_pct"].append({
            "period_label": label, "period_end": end,
            "value": nm_pct, "flagged": False,
        })

    return {
        "entity": entity,
        "dollar_scale": {"factor": dollar_factor, "label": dollar_label},
        "series": series,
    }


def _build_peer_chart_data(comparison_result):
    """
    Convert a fetch_peer_comparison result into per-company chart-ready series.

    Accepts the same comparison_result dict the SSE stream emits (with an
    embedded 'scale' key).  Does not re-fetch EDGAR.

    Periods are aligned by relative label (FY0, FY-1, …) so companies with
    different fiscal-year calendars are still comparable on the same x-axis.
    Periods are returned in chronological order (oldest first).

    Each point: {"period_label": str, "period_end": str, "value": float|None,
                 "flagged": bool}

    Margin rules mirror _build_chart_data:
      - null when either input is missing or zero
      - null when either input is flagged (flagged inputs → unreliable ratios)
    """
    companies  = comparison_result["companies"]
    n_periods  = comparison_result.get("n_periods", 3)
    scale      = comparison_result.get("scale") or pc.select_peer_scale(comparison_result)
    dollar_factor = scale["dollar_factor"]
    dollar_label  = scale["dollar_label"]

    # Chronological order: FY-(n-1), …, FY-1, FY0
    sorted_rels = ["FY-{}".format(i) for i in range(n_periods - 1, 0, -1)] + ["FY0"]

    safe_factor = dollar_factor if dollar_factor else 1

    def _scale(v):
        if v is None:
            return None
        return round(v / safe_factor, 2) if safe_factor != 1 else round(float(v), 2)

    company_series = []
    for company in companies:
        li = company["line_items"]

        def _lookup(line_item_name):
            info = li.get(line_item_name) or {}
            return {p["relative_period"]: p for p in info.get("periods", [])}

        rev_by_rel = _lookup("Revenue")
        ni_by_rel  = _lookup("Net Income")
        gp_by_rel  = _lookup("Gross Profit")

        revenue_pts        = []
        net_income_pts     = []
        gross_margin_pts   = []
        net_margin_pts     = []

        for rel in sorted_rels:
            rev_p = rev_by_rel.get(rel)
            ni_p  = ni_by_rel.get(rel)
            gp_p  = gp_by_rel.get(rel)

            period_end = (rev_p or ni_p or gp_p or {}).get("period_end", "")

            rev_val     = rev_p["value"] if rev_p else None
            rev_flagged = bool(rev_p and rev_p.get("flags"))
            ni_val      = ni_p["value"]  if ni_p  else None
            ni_flagged  = bool(ni_p  and ni_p.get("flags"))
            gp_val      = gp_p["value"]  if gp_p  else None
            gp_flagged  = bool(gp_p  and gp_p.get("flags"))

            revenue_pts.append({
                "period_label": rel, "period_end": period_end,
                "value": _scale(rev_val), "flagged": rev_flagged,
            })
            net_income_pts.append({
                "period_label": rel, "period_end": period_end,
                "value": _scale(ni_val), "flagged": ni_flagged,
            })

            if (rev_val is not None and rev_val != 0
                    and gp_val is not None
                    and not rev_flagged and not gp_flagged):
                gm_pct = round(100.0 * gp_val / rev_val, 2)
            else:
                gm_pct = None

            if (rev_val is not None and rev_val != 0
                    and ni_val is not None
                    and not rev_flagged and not ni_flagged):
                nm_pct = round(100.0 * ni_val / rev_val, 2)
            else:
                nm_pct = None

            gross_margin_pts.append({
                "period_label": rel, "period_end": period_end,
                "value": gm_pct, "flagged": False,
            })
            net_margin_pts.append({
                "period_label": rel, "period_end": period_end,
                "value": nm_pct, "flagged": False,
            })

        company_series.append({
            "name": company["name"],
            "cik":  company["cik"],
            "series": {
                "revenue":          revenue_pts,
                "net_income":       net_income_pts,
                "gross_margin_pct": gross_margin_pts,
                "net_margin_pct":   net_margin_pts,
            },
        })

    return {
        "periods":      sorted_rels,
        "dollar_scale": {"factor": dollar_factor, "label": dollar_label},
        "companies":    company_series,
    }


def _xbrl_write_csv(filepath, entity, columns, rows, period_type):
    import csv
    dollar_factor, dollar_label = _detect_dollar_scale(rows, columns)
    share_factor, share_label = _detect_share_scale(rows, columns)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["# {} -- XBRL Financial Data ({})".format(entity, period_type)])
        writer.writerow(["Line Item", "Source Tag"] + [c["label"] for c in columns] + ["Flags"])
        for row in rows:
            line_item = row["line_item"]
            if line_item in _DOLLAR_LINE_ITEMS:
                factor, suffix = dollar_factor, " ({})".format(dollar_label)
            elif line_item in _SHARE_LINE_ITEMS:
                factor, suffix = share_factor, " ({})".format(share_label)
            else:
                factor, suffix = 1, ""
            label = line_item + suffix
            cells_out = []
            all_row_flags = []
            for col in columns:
                cell = row["cells"].get(col["key"])
                val = cell.get("value") if cell else None
                if val is None:
                    cells_out.append("Not reported")
                elif line_item in _EPS_LINE_ITEMS:
                    cells_out.append("{:.2f}".format(val))
                else:
                    scaled = val / factor if factor != 1 else val
                    cells_out.append("{:,.0f}".format(scaled))
                if cell:
                    all_row_flags.extend(f["msg"] for f in cell.get("flags", []))
            flags_str = "; ".join(sorted(set(all_row_flags))) if all_row_flags else ""
            # The Source Tag column names every tag the row's values came from,
            # in period order, because a stitched row can span more than one.
            source = row.get("tag_summary") or row.get("tag_used") or ""
            writer.writerow([label, source] + cells_out + [flags_str])


def _xbrl_add_chart_sheet(wb, entity, columns, rows, dollar_factor, dollar_label):
    """
    Add a 'Charts' sheet with two native openpyxl LineCharts.

    Chart 1: Revenue and Net Income (scaled, dollar_label on y-axis).
    Chart 2: Gross Margin % and Net Margin % (% on y-axis).

    Missing data → empty cell → Excel renders as a gap (never as zero).
    Per-point flagged styling is not available in openpyxl LineChart; a
    text note is written instead when flagged values are present.
    """
    try:
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        return  # skip silently if openpyxl chart objects not available

    chart_data = _build_chart_data(entity, columns, rows, dollar_factor, dollar_label)
    s      = chart_data["series"]
    n_cols = len(columns)

    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = "16A34A"

    # ------------------------------------------------------------------
    # Helper data table — column-oriented (each period is a column)
    # Row 1: period labels
    # Rows 2-5: series data; None written as empty cell (Excel gap)
    # ------------------------------------------------------------------
    ws.cell(1, 1, "Period")
    for ci, col in enumerate(columns, start=2):
        ws.cell(1, ci, col["label"])

    series_layout = [
        (2, "Revenue ({})".format(dollar_label),    s["revenue"]),
        (3, "Net Income ({})".format(dollar_label),  s["net_income"]),
        (4, "Gross Margin %",                        s["gross_margin_pct"]),
        (5, "Net Margin %",                          s["net_margin_pct"]),
    ]
    for row_idx, label, pts in series_layout:
        ws.cell(row_idx, 1, label)
        for ci, pt in enumerate(pts, start=2):
            if pt["value"] is not None:
                ws.cell(row_idx, ci, pt["value"])

    # Flag note
    has_flags = any(
        pt["flagged"]
        for key in ("revenue", "net_income")
        for pt in s[key]
    )
    note_row = 6
    if has_flags:
        ws.cell(note_row, 1,
            "Note: one or more plotted values are flagged for review — "
            "see 'Financial Data' sheet (yellow highlighted cells) for details."
        )
        note_row = 7

    # ------------------------------------------------------------------
    # Chart 1: Revenue & Net Income
    # ------------------------------------------------------------------
    c1 = LineChart()
    c1.title        = "{} — Revenue & Net Income".format(entity)
    c1.y_axis.title = dollar_label
    c1.x_axis.title = "Period"
    c1.style        = 10
    c1.width        = 22
    c1.height       = 14

    cats = Reference(ws, min_col=2, max_col=1 + n_cols, min_row=1, max_row=1)
    # Rows 2-3; col 1 = series title via titles_from_data=True
    data1 = Reference(ws, min_col=1, max_col=1 + n_cols, min_row=2, max_row=3)
    c1.add_data(data1, from_rows=True, titles_from_data=True)
    c1.set_categories(cats)

    ws.add_chart(c1, "A{}".format(note_row + 1))

    # ------------------------------------------------------------------
    # Chart 2: Gross Margin % and Net Margin %
    # ------------------------------------------------------------------
    c2 = LineChart()
    c2.title        = "{} — Gross Margin % and Net Margin %".format(entity)
    c2.y_axis.title = "%"
    c2.x_axis.title = "Period"
    c2.style        = 10
    c2.width        = 22
    c2.height       = 14

    data2 = Reference(ws, min_col=1, max_col=1 + n_cols, min_row=4, max_row=5)
    c2.add_data(data2, from_rows=True, titles_from_data=True)
    c2.set_categories(cats)

    ws.add_chart(c2, "A{}".format(note_row + 26))


def _peer_add_chart_sheet(wb, comparison_result):
    """
    Add a 'Charts' sheet with four native openpyxl LineCharts.

    Chart 1: Revenue per company.
    Chart 2: Net Income per company.
    Chart 3: Gross Margin % per company.
    Chart 4: Net Margin % per company.
    X-axis uses relative period labels in chronological order (oldest → newest).
    Missing / derived-unavailable data → empty cell → Excel gap per company.
    """
    try:
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        return

    # Use _build_peer_chart_data so margins are computed consistently
    chart_data    = _build_peer_chart_data(comparison_result)
    companies_cd  = chart_data["companies"]   # [{name, cik, series: {revenue, net_income, …}}]
    sorted_rels   = chart_data["periods"]     # chronological: FY-(n-1) … FY0
    dollar_label  = chart_data["dollar_scale"]["label"]
    n_companies   = len(companies_cd)
    n_periods     = len(sorted_rels)

    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = "16A34A"

    def _write_block(start_row, metric_label, series_key):
        """Write one metric block; returns the row immediately after."""
        ws.cell(start_row, 1, metric_label)
        for ci, rel in enumerate(sorted_rels, start=2):
            ws.cell(start_row, ci, rel)
        for ri, co in enumerate(companies_cd, start=1):
            ws.cell(start_row + ri, 1, co["name"])
            for ci, pt in enumerate(co["series"][series_key], start=2):
                if pt["value"] is not None:
                    ws.cell(start_row + ri, ci, pt["value"])
        return start_row + n_companies + 1

    metrics = [
        ("Revenue ({})".format(dollar_label),    "revenue",          dollar_label),
        ("Net Income ({})".format(dollar_label),  "net_income",       dollar_label),
        ("Gross Margin %",                        "gross_margin_pct", "%"),
        ("Net Margin %",                          "net_margin_pct",   "%"),
    ]

    starts = []
    cur_row = 1
    for label, key, _ in metrics:
        starts.append(cur_row)
        cur_row = _write_block(cur_row, label, key)
        cur_row += 1   # blank row between blocks

    note_row = cur_row
    ws.cell(note_row, 1,
        "Note: flagged values appear highlighted on the Comparison sheet. "
        "Per-point chart markers are not supported in native Excel LineCharts."
    )
    chart_start = note_row + 2

    names_short = " vs ".join(co["name"] for co in companies_cd[:3])
    if n_companies > 3:
        names_short += " & {} more".format(n_companies - 3)

    for idx, ((label, key, y_label), start) in enumerate(zip(metrics, starts)):
        c = LineChart()
        c.title        = "{} — {}".format(label, names_short)
        c.y_axis.title = y_label
        c.x_axis.title = "Period (oldest → newest)"
        c.style        = 10
        c.width        = 22
        c.height       = 14

        cats = Reference(ws, min_col=2, max_col=1 + n_periods,
                         min_row=start, max_row=start)
        data = Reference(ws, min_col=1, max_col=1 + n_periods,
                         min_row=start + 1, max_row=start + n_companies)
        c.add_data(data, from_rows=True, titles_from_data=True)
        c.set_categories(cats)
        ws.add_chart(c, "A{}".format(chart_start + idx * 26))


_PROVENANCE_SHEET_HEADERS = (
    ("Line Item", 28), ("Period", 12), ("Source", 11), ("XBRL Tag", 46),
    ("Filed", 12), ("Accession", 22), ("Notes", 80),
)


def _xbrl_add_source_tags_sheet(wb, columns, rows, BF):
    """Add a 'Source Tags' sheet holding one line per value, not per row.

    A row of a stitched series can span several tags, so a single tag_used
    column on the data sheet cannot say where any particular number came from.
    This sheet answers that per period: which of the three states the value is
    in, which tag reported it, when it was filed and under which accession, the
    formula behind a derived value, and where to go look for a missing one.
    Tag seams are called out on the first period of the new tag.
    """
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("Source Tags")
    ws.sheet_properties.tabColor = "888888"

    hdr_font = Font(name=BF, color="FFFFFF", bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="003366")
    base_font = Font(name=BF, size=11)
    tag_font = Font(name=BF, color="888888", size=10)
    missing_font = Font(name=BF, color="AAAAAA", italic=True, size=11)
    note_font = Font(name=BF, color="666666", size=10)
    seam_font = Font(name=BF, color="856404", size=10)

    for ci, (text, width) in enumerate(_PROVENANCE_SHEET_HEADERS, start=1):
        cell = ws.cell(1, ci, text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        ws.column_dimensions[cell.column_letter].width = width

    out_row = 2
    for row in rows:
        for col in columns:
            cell = row["cells"].get(col["key"]) or {}
            prov = cell.get("provenance") or {}
            state = prov.get("state") or xbrl.PROVENANCE_MISSING

            note = ""
            if state == xbrl.PROVENANCE_DERIVED:
                note = "Derived: {}".format(prov.get("formula", ""))
            elif state == xbrl.PROVENANCE_MISSING:
                note = prov.get("message", "")
            else:
                seam = next((f for f in cell.get("flags", [])
                             if f.get("type") == xbrl.FLAG_TAG_TRANSITION), None)
                if seam:
                    note = seam.get("msg", "")

            ws.cell(out_row, 1, row["line_item"]).font = base_font
            ws.cell(out_row, 2, col["label"]).font = base_font
            ws.cell(out_row, 3, state).font = (
                missing_font if state == xbrl.PROVENANCE_MISSING else base_font)
            ws.cell(out_row, 4, prov.get("tag") or "").font = tag_font
            ws.cell(out_row, 5, prov.get("filed") or "").font = tag_font
            ws.cell(out_row, 6, prov.get("accession") or "").font = tag_font
            ws.cell(out_row, 7, note).font = seam_font if note.startswith(
                row["line_item"] + " switches") else note_font
            out_row += 1

    ws.freeze_panes = "A2"
    ws.cell(out_row + 1, 1,
            "One line per value. reported = the filer tagged it; derived = "
            "Edgardly computed it from reported values, formula in Notes; "
            "missing = nobody tagged it, Notes says where to look."
            ).font = Font(name=BF, color="666666", italic=True, size=10)
    return ws


def _xbrl_write_xlsx(filepath, entity, columns, rows, period_type):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    dollar_factor, dollar_label = _detect_dollar_scale(rows, columns)
    share_factor, share_label = _detect_share_scale(rows, columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    ws.sheet_properties.tabColor = "003366"

    BF = "Calibri"
    hdr_fill = PatternFill("solid", fgColor="003366")
    hdr_font = Font(name=BF, color="FFFFFF", bold=True, size=11)
    flag_fill = PatternFill("solid", fgColor="FFF3CD")
    tag_font = Font(name=BF, color="888888", size=10)
    flag_text_font = Font(name=BF, color="856404", size=11)
    base_font = Font(name=BF, size=11)
    missing_font = Font(name=BF, color="AAAAAA", italic=True, size=11)
    extracted_font = Font(name=BF, color="0066CC", size=11)   # blue: direct XBRL values
    calc_font = Font(name=BF, size=11)                          # black: computed/derived values
    flag_data_font = Font(name=BF, color="CC0000", size=11)    # red: flagged data cells
    _border_thin   = Border(bottom=Side(style="thin"))
    _border_double = Border(bottom=Side(style="double"))

    # Row 1: title
    tc = ws.cell(1, 1, "{} -- XBRL Financial Data ({})".format(entity, period_type.title()))
    tc.font = Font(name=BF, bold=True, size=12)

    hrow = 3
    col_widths = {}

    # Header row
    for ci, (text, min_w) in enumerate([("Line Item", 28), ("Source Tag", 15)], start=1):
        c = ws.cell(hrow, ci, text)
        c.font = hdr_font
        c.fill = hdr_fill
        col_widths[ci] = max(min_w, len(text))

    for ci, col in enumerate(columns, start=3):
        c = ws.cell(hrow, ci, col["label"])
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="right")
        col_widths[ci] = len(col["label"])

    flags_col = len(columns) + 3
    c = ws.cell(hrow, flags_col, "Flags")
    c.font = hdr_font
    c.fill = hdr_fill
    col_widths[flags_col] = 12

    # Data rows
    for ri, row in enumerate(rows, start=hrow + 1):
        line_item = row["line_item"]
        is_dollar = line_item in _DOLLAR_LINE_ITEMS
        is_eps = line_item in _EPS_LINE_ITEMS
        is_shares = line_item in _SHARE_LINE_ITEMS

        if is_dollar:
            factor, suffix, num_fmt = dollar_factor, " ({})".format(dollar_label), _XLSX_FMT_DOLLAR
        elif is_shares:
            factor, suffix, num_fmt = share_factor, " ({})".format(share_label), _XLSX_FMT_SHARES
        elif is_eps:
            factor, suffix, num_fmt = 1, "", _XLSX_FMT_EPS
        else:
            factor, suffix, num_fmt = 1, "", _XLSX_FMT_DOLLAR

        label = line_item + suffix
        c = ws.cell(ri, 1, label)
        c.font = base_font
        col_widths[1] = min(42, max(col_widths.get(1, 0), len(label)))

        tag_val = row.get("tag_summary") or row.get("tag_used") or ""
        c = ws.cell(ri, 2, tag_val)
        c.font = tag_font
        col_widths[2] = min(52, max(col_widths.get(2, 0), len(tag_val)))

        all_row_flags = []

        for ci, col in enumerate(columns, start=3):
            cell = row["cells"].get(col["key"])
            c = ws.cell(ri, ci)
            raw_val = cell.get("value") if cell else None
            if raw_val is None:
                c.value = "Not reported"
                c.font = missing_font
                c.alignment = Alignment(horizontal="right")
            else:
                scaled = raw_val / factor if factor != 1 else raw_val
                # Write as int when the scaled value has no fractional part (except EPS)
                if is_eps:
                    c.value = float(scaled)
                elif isinstance(scaled, float) and scaled == int(scaled):
                    c.value = int(scaled)
                else:
                    c.value = float(scaled)
                c.number_format = num_fmt
                abs_s = abs(scaled)
                dstr = "{:,.2f}".format(abs_s) if is_eps else "{:,.0f}".format(abs_s)
                col_widths[ci] = max(col_widths.get(ci, 0), len(dstr) + 3)
                if cell.get("flags"):
                    c.fill = flag_fill
                    c.font = flag_data_font
                    all_row_flags.extend(f["msg"] for f in cell["flags"])
                else:
                    # Blue for a value the filer reported, black for one
                    # Edgardly computed. The split is per value, because one
                    # row can hold both.
                    state = (cell.get("provenance") or {}).get("state")
                    c.font = calc_font if state == xbrl.PROVENANCE_DERIVED else extracted_font
                c.alignment = Alignment(horizontal="right")

        # Bottom border: single under subtotals, double under final total
        if line_item in _SINGLE_BORDER_ROW_ITEMS:
            _rb = _border_thin
        elif line_item in _DOUBLE_BORDER_ROW_ITEMS:
            _rb = _border_double
        else:
            _rb = None
        if _rb:
            for _bc in range(1, flags_col + 1):
                ws.cell(ri, _bc).border = _rb

        flags_str = "; ".join(sorted(set(all_row_flags))) if all_row_flags else ""
        if flags_str:
            c2 = ws.cell(ri, flags_col, flags_str)
            c2.font = flag_text_font
            col_widths[flags_col] = max(col_widths.get(flags_col, 0), min(len(flags_str), 60))

    # Apply auto-fit column widths
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w + 2

    # Freeze title + header rows AND left label column so scrolling works in both axes
    ws.freeze_panes = "B4"

    # Sanity-check rows (one blank separator row, then BS and GP reconciliation)
    _row_map = {row["line_item"]: (hrow + 1 + ri) for ri, row in enumerate(rows)}
    _data_col_idxs = list(range(3, 3 + len(columns)))
    _legend_row = _write_xlsx_sanity_checks(
        ws, hrow + len(rows) + 2, _row_map, _data_col_idxs, dollar_label, BF
    )

    ws.cell(_legend_row, 1,
            "Blue = reported by the filer.  Black = derived by Edgardly, formula on the "
            "Source Tags sheet.  Grey \"Not reported\" = nobody tagged it, and the Source "
            "Tags sheet says which statement of which filing to check."
            ).font = Font(name=BF, color="666666", italic=True, size=10)

    _xbrl_add_source_tags_sheet(wb, columns, rows, BF)
    _xbrl_add_chart_sheet(wb, entity, columns, rows, dollar_factor, dollar_label)
    wb.save(filepath)


@app.route("/")
def index():
    # The page needs to know which line items exist and what units each carries.
    # That is the registry's job, so it is injected here rather than typed out a
    # second time in the template, where nothing would keep the two in step.
    return render_template(
        "index.html",
        line_item_classification=line_items.classification_for_client(),
    )


@app.route("/api/search")
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    results = edgar_api.search_companies(q)
    return jsonify(results)


@app.route("/api/filings")
def api_filings():
    cik = request.args.get("cik", "").strip()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    forms_param = request.args.get("forms", "").strip()
    form_types = [f.strip() for f in forms_param.split(",") if f.strip()] if forms_param else ["10-K"]
    if not cik or not start or not end:
        return jsonify({"error": "cik, start, and end are required"}), 400
    try:
        filings = edgar_api.get_filings(cik, start, end, form_types=form_types)
        return jsonify(filings)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download", methods=["POST"])
def api_download():
    data = request.get_json(silent=True)
    required = {"cik", "accession_number", "company_name", "filing_date", "fiscal_year_end", "form_type"}
    if not data or not required.issubset(data.keys()):
        missing = required - set(data or {})
        return jsonify({"error": f"Missing required fields: {', '.join(missing)}"}), 400
    document_url = data.get("document_url", "")
    if document_url and not document_url.startswith("https://www.sec.gov/"):
        return jsonify({"error": "Invalid document URL"}), 400
    fmt = data.get("format", "html")
    if fmt not in ("html", "pdf", "both"):
        return jsonify({"error": "format must be 'html', 'pdf', or 'both'"}), 400
    try:
        result = edgar_api.download_filing(
            cik=data["cik"],
            accession_number=data["accession_number"],
            company_name=data["company_name"],
            filing_date=data["filing_date"],
            fiscal_year_end=data["fiscal_year_end"],
            form_type=data["form_type"],
            downloads_dir=DOWNLOADS_DIR,
            url=document_url or None,
            fmt=fmt,
        )
        primary = result['primary']
        return jsonify({
            "status": "ok",
            "path": primary,
            "folder": os.path.dirname(primary),
            "pdf_fallback": result.get('pdf_fallback', False),
        })
    except edgar_api.FilingNotFoundError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download-batch", methods=["POST"])
def api_download_batch():
    body = request.get_json(silent=True)
    if not isinstance(body, dict):
        return jsonify({"error": "Expected a JSON object with 'filings' array"}), 400
    filings_data = body.get('filings', [])
    if not isinstance(filings_data, list) or not filings_data:
        return jsonify({"error": "No filings provided"}), 400
    fmt = body.get('format', 'html')
    if fmt not in ('html', 'pdf', 'both'):
        return jsonify({"error": "format must be 'html', 'pdf', or 'both'"}), 400
    required = {"cik", "accession_number", "company_name", "filing_date", "fiscal_year_end", "form_type"}
    for i, item in enumerate(filings_data):
        missing = required - set(item.keys())
        if missing:
            return jsonify({"error": f"Item {i} missing fields: {', '.join(missing)}"}), 400
        url = item.get("document_url", "")
        if url and not url.startswith("https://www.sec.gov/"):
            return jsonify({"error": f"Item {i} has invalid document URL"}), 400
    results = edgar_api.download_filings_batch(filings_data, DOWNLOADS_DIR, fmt=fmt)
    folder = None
    for r in results:
        if r.get('status') == 'ok' and r.get('path'):
            folder = os.path.dirname(os.path.dirname(r['path']))
            break
    return jsonify({"results": results, "folder": folder})


@app.route("/api/downloads")
def api_downloads():
    return jsonify(edgar_api.list_downloads(DOWNLOADS_DIR))


@app.route("/downloads/<path:filepath>")
def serve_download(filepath):
    return send_from_directory(DOWNLOADS_DIR, filepath)


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No data provided"}), 400
    for field in ("cik", "start", "end"):
        if not data.get(field):
            return jsonify({"error": f"Missing required field: {field}"}), 400
    fmt = data.get("format", "xlsx")
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "format must be 'csv' or 'xlsx'"}), 400
    forms_param = data.get("forms", "")
    form_types = [f.strip() for f in forms_param.split(",") if f.strip()] if forms_param else ["10-K"]
    try:
        filepath = edgar_api.export_filings(
            cik=data["cik"],
            start_date=data["start"],
            end_date=data["end"],
            form_types=form_types,
            fmt=fmt,
            company_name=data.get("company_name", ""),
            ticker=data.get("ticker", ""),
            exports_dir=EXPORTS_DIR,
        )
        filename = os.path.basename(filepath)
        rel_path = os.path.relpath(filepath, EXPORTS_DIR).replace('\\', '/')
        return jsonify({
            "status": "ok",
            "path": filepath,
            "filename": filename,
            "folder": os.path.dirname(filepath),
            "download_url": "/exports/" + rel_path,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/open-folder", methods=["POST"])
def api_open_folder():
    data = request.get_json(silent=True) or {}
    folder = data.get("folder", EXPORTS_DIR)
    folder = os.path.normpath(folder)
    allowed = (os.path.normpath(EXPORTS_DIR), os.path.normpath(DOWNLOADS_DIR))
    if not any(folder.startswith(root) for root in allowed):
        folder = EXPORTS_DIR
    try:
        os.startfile(folder)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/exports/<path:filepath>")
def serve_export(filepath):
    return send_from_directory(EXPORTS_DIR, filepath)


@app.route("/api/xbrl/extract", methods=["POST"])
def api_xbrl_extract():
    data = request.get_json(silent=True) or {}
    cik = str(data.get("cik", "")).strip()
    if not cik:
        return jsonify({"error": "cik is required"}), 400
    try:
        start_year = int(data.get("start_year", 2015))
        end_year = int(data.get("end_year", datetime.date.today().year))
    except (ValueError, TypeError):
        return jsonify({"error": "start_year and end_year must be integers"}), 400
    period_type = data.get("period_type", "annual")
    if period_type not in ("annual", "quarterly"):
        return jsonify({"error": "period_type must be 'annual' or 'quarterly'"}), 400
    try:
        entity, columns, rows, scope = _build_xbrl_result(
            cik, start_year, end_year, period_type)
        return jsonify({
            "entity": entity, "columns": columns, "rows": rows, "scope": scope,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/xbrl/chart-data", methods=["POST"])
def api_xbrl_chart_data():
    """Return chart-ready series for Revenue, Net Income, Gross Margin %,
    and Net Margin % using the same extraction logic as /api/xbrl/extract."""
    data = request.get_json(silent=True) or {}
    cik = str(data.get("cik", "")).strip()
    if not cik:
        return jsonify({"error": "cik is required"}), 400
    try:
        start_year = int(data.get("start_year", 2015))
        end_year   = int(data.get("end_year", datetime.date.today().year))
    except (ValueError, TypeError):
        return jsonify({"error": "start_year and end_year must be integers"}), 400
    period_type = data.get("period_type", "annual")
    if period_type not in ("annual", "quarterly"):
        return jsonify({"error": "period_type must be 'annual' or 'quarterly'"}), 400
    try:
        entity, columns, rows, _scope = _build_xbrl_result(
            cik, start_year, end_year, period_type)
        dollar_factor, dollar_label = _detect_dollar_scale(rows, columns)
        return jsonify(_build_chart_data(entity, columns, rows, dollar_factor, dollar_label))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/xbrl/export", methods=["POST"])
def api_xbrl_export():
    data = request.get_json(silent=True) or {}
    cik = str(data.get("cik", "")).strip()
    if not cik:
        return jsonify({"error": "cik is required"}), 400
    try:
        start_year = int(data.get("start_year", 2015))
        end_year = int(data.get("end_year", datetime.date.today().year))
    except (ValueError, TypeError):
        return jsonify({"error": "start_year and end_year must be integers"}), 400
    period_type = data.get("period_type", "annual")
    if period_type not in ("annual", "quarterly"):
        return jsonify({"error": "period_type must be 'annual' or 'quarterly'"}), 400
    fmt = data.get("format", "xlsx")
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "format must be 'csv' or 'xlsx'"}), 400
    try:
        entity, columns, rows, _scope = _build_xbrl_result(
            cik, start_year, end_year, period_type)
        safe = "".join(c if c.isalnum() or c in " -_" else "" for c in entity)[:40].strip()
        company_folder = os.path.join(EXPORTS_DIR, safe.replace(" ", "_") or cik)
        os.makedirs(company_folder, exist_ok=True)
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H%M")
        filename = "{}_XBRL_{}_{}_{}{}".format(
            safe.replace(" ", "_") or cik,
            period_type.title(),
            date_str,
            time_str,
            "." + fmt,
        )
        filepath = os.path.join(company_folder, filename)
        if fmt == "csv":
            _xbrl_write_csv(filepath, entity, columns, rows, period_type)
        else:
            _xbrl_write_xlsx(filepath, entity, columns, rows, period_type)
        rel_path = os.path.relpath(filepath, EXPORTS_DIR).replace("\\", "/")
        return jsonify({
            "status": "ok",
            "filename": filename,
            "folder": company_folder,
            "download_url": "/exports/" + rel_path,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Three-statement scaffold (V2_PLAN 2.3)
#
# Composition only. What a scaffold contains is decided in
# app/scaffold/three_statement.py and what the file looks like in
# app/scaffold/excel.py; this endpoint fetches the two things those need, calls
# them in order, and turns a refusal into an HTTP response. No finance and no
# openpyxl belong here, and the day a scaffold rule changes, nothing in this
# file should have to.
# ---------------------------------------------------------------------------

SCAFFOLD_HISTORY_YEARS = 5
SCAFFOLD_FORECAST_YEARS = 3
SCAFFOLD_MAX_YEARS = 20


def _scaffold_filename(entity, cik):
    safe = "".join(c if c.isalnum() or c in " -_" else "" for c in entity)[:40].strip()
    stem = safe.replace(" ", "_") or str(cik)
    now = datetime.datetime.now()
    return stem, "{}_3Statement_{}_{}.xlsx".format(
        stem, now.strftime("%Y-%m-%d"), now.strftime("%H%M"))


@app.route("/api/scaffold/three-statement", methods=["POST"])
def api_scaffold_three_statement():
    """Build one filer's three-statement scaffold workbook.

    A filer the scope gate refuses gets 422 and the gate's own message, which is
    the same sentence the XBRL view already shows under the table. The refusal
    is the whole response: no half-built workbook is written and no file is left
    behind for something else to pick up by mistake.
    """
    data = request.get_json(silent=True) or {}
    cik = str(data.get("cik", "")).strip()
    if not cik:
        return jsonify({"error": "cik is required"}), 400
    try:
        years = int(data.get("years", SCAFFOLD_HISTORY_YEARS))
        forecast_years = int(data.get("forecast_years", SCAFFOLD_FORECAST_YEARS))
    except (ValueError, TypeError):
        return jsonify({"error": "years and forecast_years must be integers"}), 400
    if not 1 <= years <= SCAFFOLD_MAX_YEARS:
        return jsonify({"error": "years must be between 1 and {}".format(
            SCAFFOLD_MAX_YEARS)}), 400
    if not 0 <= forecast_years <= SCAFFOLD_MAX_YEARS:
        return jsonify({"error": "forecast_years must be between 0 and {}".format(
            SCAFFOLD_MAX_YEARS)}), 400
    fmt = data.get("format", "xlsx")
    if fmt != "xlsx":
        # A scaffold is linked formulas across seven sheets. CSV would flatten
        # exactly the thing it exists to build, so it is refused rather than
        # silently written as values.
        return jsonify({"error": "format must be 'xlsx'; a scaffold is linked "
                                 "formulas and cannot be written as CSV"}), 400

    try:
        facts = xbrl.fetch_company_facts(cik)
        sic, sic_lookup = _company_sic(cik)
        spec = scaffold_model.build_model(cik, facts, sic, history_years=years,
                                          forecast_years=forecast_years)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    scope = {"in_scope": spec.scope.in_scope, "reason": spec.scope.reason,
             "message": spec.scope.message,
             "detail": dict(spec.scope.detail, sic_lookup=sic_lookup)}
    if not spec.scope.in_scope:
        return jsonify({"error": spec.scope.message, "scope": scope,
                        "entity": spec.entity}), 422
    if not spec.rows:
        message = (spec.flags[0]["message"] if spec.flags
                   else "There is no model to build for this filer.")
        return jsonify({"error": message, "scope": scope,
                        "entity": spec.entity}), 422

    try:
        stem, filename = _scaffold_filename(spec.entity, cik)
        folder = os.path.join(EXPORTS_DIR, stem)
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, filename)
        scaffold_excel.write_workbook(spec, filepath)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    rel_path = os.path.relpath(filepath, EXPORTS_DIR).replace("\\", "/")
    return jsonify({
        "status": "ok",
        "entity": spec.entity,
        "filename": filename,
        "folder": folder,
        "download_url": "/exports/" + rel_path,
        "scope": scope,
        "historical": [p.label for p in scaffold_model.historical_periods(spec)],
        "forecast": [p.label for p in scaffold_model.forecast_periods(spec)],
        "flags": [{"flag_type": f["flag_type"], "message": f["message"]}
                  for f in scaffold_model.summarised_flags(spec)],
    })


# ---------------------------------------------------------------------------
# Peer comparison -- Stage 3 (stream) and Stage 4 (export)
# ---------------------------------------------------------------------------

def _peer_write_xlsx(filepath, comparison_result):
    """Write a peer comparison result to an Excel workbook with two sheets:
      'Comparison' -- the scaled table, flagged cells highlighted, N/A for missing.
      'Source Tags' -- which XBRL tag was used for each company / line item.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    companies  = comparison_result["companies"]
    line_items = comparison_result["line_items"]
    n_periods  = comparison_result["n_periods"]
    scale      = comparison_result.get("scale", pc.select_peer_scale(comparison_result))

    dollar_factor = scale["dollar_factor"]
    dollar_label  = scale["dollar_label"]
    share_factor  = scale["share_factor"]
    share_label   = scale["share_label"]

    # ---- constants ----
    BF         = "Calibri"
    HDR_FILL   = PatternFill("solid", fgColor="003366")
    GRP_FILL   = PatternFill("solid", fgColor="1A5276")   # slightly lighter for company rows
    HDR_FONT   = Font(name=BF, color="FFFFFF", bold=True, size=11)
    GRP_FONT   = Font(name=BF, color="FFFFFF", bold=True, size=11)
    BASE_FONT  = Font(name=BF, size=11)
    MISS_FONT  = Font(name=BF, color="AAAAAA", italic=True, size=11)
    TAG_FONT   = Font(name=BF, color="888888", size=10)
    FLAG_FILL  = PatternFill("solid", fgColor="FFF3CD")
    FLAG_FONT  = Font(name=BF, color="856404", size=11)   # kept for flag-details text
    FLAG_DATA_FONT = Font(name=BF, color="CC0000", size=11)  # red font on flagged data cells
    EXTRACTED_FONT = Font(name=BF, color="0066CC", size=11)  # blue: XBRL-sourced values
    CALC_FONT  = Font(name=BF, size=11)                      # black: derived values
    FLAG_BORDER = Border(
        left=Side(style="medium", color="DC3545"),
    )
    _border_thin   = Border(bottom=Side(style="thin"))
    _border_double = Border(bottom=Side(style="double"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.sheet_properties.tabColor = "003366"

    # ---- Row 1: title ----
    company_names = " vs ".join(c["name"] for c in companies)
    title = "Peer Comparison: {}  |  Scale: {} / {}".format(
        company_names, dollar_label, share_label + " shares"
    )
    tc = ws.cell(1, 1, title)
    tc.font = Font(name=BF, bold=True, size=12)

    # ---- Column layout ----
    # Col 1: Line Item label
    # Then for each company: n_periods columns
    first_val_col = 2   # first data column

    def _col(company_idx, period_idx):
        return first_val_col + company_idx * n_periods + period_idx

    # ---- Row 2: company group headers ----
    c = ws.cell(2, 1, "Line Item")
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center")

    for ci, company in enumerate(companies):
        col_start = _col(ci, 0)
        col_end   = _col(ci, n_periods - 1)
        ws.cell(2, col_start, company["name"]).fill = GRP_FILL
        ws.cell(2, col_start).font  = GRP_FONT
        ws.cell(2, col_start).alignment = Alignment(horizontal="center", vertical="center")
        if n_periods > 1:
            ws.merge_cells(
                start_row=2, start_column=col_start,
                end_row=2,   end_column=col_end,
            )

    # ---- Row 3: period label headers ----
    ws.cell(3, 1)  # empty corner
    ws.cell(3, 1).fill = HDR_FILL

    for ci, company in enumerate(companies):
        company_data = company["line_items"].get(line_items[0]) if line_items else None
        for pi in range(n_periods):
            col = _col(ci, pi)
            rel_label = "FY0" if pi == 0 else "FY-{}".format(pi)
            c = ws.cell(3, col, rel_label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "B4"

    col_widths = {1: 28}

    # ---- Row 4+: data ----
    for ri, li in enumerate(line_items, start=4):
        is_dollar = li in pc.DOLLAR_LINE_ITEMS
        is_eps    = li in pc.EPS_LINE_ITEMS
        is_shares = li in pc.SHARE_LINE_ITEMS

        if is_dollar:
            factor, suffix, num_fmt = dollar_factor, " ({})".format(dollar_label), _XLSX_FMT_DOLLAR
        elif is_shares:
            factor, suffix, num_fmt = share_factor, " ({})".format(share_label), _XLSX_FMT_SHARES
        elif is_eps:
            factor, suffix, num_fmt = 1, "", _XLSX_FMT_EPS
        else:
            factor, suffix, num_fmt = 1, "", _XLSX_FMT_DOLLAR

        label = li + suffix
        lc = ws.cell(ri, 1, label)
        lc.font = BASE_FONT
        col_widths[1] = min(42, max(col_widths.get(1, 0), len(label)))

        total_data_cols = first_val_col + len(companies) * n_periods - 1

        for ci, company in enumerate(companies):
            item_info = company["line_items"].get(li) or {}
            periods   = item_info.get("periods", [])

            for pi in range(n_periods):
                col = _col(ci, pi)
                c   = ws.cell(ri, col)
                period = periods[pi] if pi < len(periods) else None
                value  = period["value"] if period else None

                if value is None:
                    c.value = "N/A"
                    c.font  = MISS_FONT
                    c.alignment = Alignment(horizontal="right")
                else:
                    scaled = value / factor if factor != 1 else value
                    if is_eps:
                        c.value = float(scaled)
                    elif isinstance(scaled, float) and scaled == int(scaled):
                        c.value = int(scaled)
                    else:
                        c.value = float(scaled)
                    c.number_format = num_fmt
                    c.alignment = Alignment(horizontal="right")
                    abs_s = abs(scaled)
                    dstr = "{:,.2f}".format(abs_s) if is_eps else "{:,.0f}".format(abs_s)
                    col_widths[col] = max(col_widths.get(col, len(rel_label) + 2), len(dstr) + 3)

                    if period and period.get("flags"):
                        c.fill   = FLAG_FILL
                        c.border = FLAG_BORDER
                        c.font   = FLAG_DATA_FONT
                        c.number_format = num_fmt
                    else:
                        # Blue reported, black derived, per value.
                        state = (period or {}).get("provenance", {}).get("state")
                        c.font = (CALC_FONT if state == xbrl.PROVENANCE_DERIVED
                                  else EXTRACTED_FONT)

        # Bottom border: single under subtotals, double under Net Income
        if li in _SINGLE_BORDER_ROW_ITEMS:
            _rb = _border_thin
        elif li in _DOUBLE_BORDER_ROW_ITEMS:
            _rb = _border_double
        else:
            _rb = None
        if _rb:
            for _bc in range(1, total_data_cols + 1):
                ws.cell(ri, _bc).border = _rb

    # Auto-fit
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w + 2

    # Ensure every value column has at least a minimum width
    total_cols = first_val_col + len(companies) * n_periods - 1
    for col_idx in range(first_val_col, total_cols + 1):
        if col_idx not in col_widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

    # Sanity-check rows (one blank separator row, then BS and GP reconciliation)
    _row_map = {li: (4 + idx) for idx, li in enumerate(line_items)}
    _data_col_idxs = [_col(ci, pi)
                      for ci in range(len(companies))
                      for pi in range(n_periods)]
    next_row = _write_xlsx_sanity_checks(
        ws, 4 + len(line_items) + 2, _row_map, _data_col_idxs, dollar_label, BF
    )

    # Validation-flag legend (after sanity checks)
    ws.cell(next_row, 1,
            "Yellow cells with red left border have validation flags. "
            "Explanations listed below (if any). See Source Tags sheet for XBRL tag details."
    ).font = Font(name=BF, color="666666", italic=True, size=10)

    # ---- Flag Details section ----
    flag_entries = []
    for company in companies:
        for li in line_items:
            item_info = company["line_items"].get(li) or {}
            for period in item_info.get("periods", []):
                for f in period.get("flags", []):
                    msg = f.get("message") or f.get("msg") or ""
                    if msg:
                        flag_entries.append({
                            "company": company["name"],
                            "period":  period.get("relative_period", ""),
                            "li":      li,
                            "msg":     msg,
                        })

    if flag_entries:
        fd_start = next_row + 2
        sh = ws.cell(fd_start, 1, "Flag Details")
        sh.font = Font(name=BF, bold=True, size=11, color="856404")
        fd_hdr = fd_start + 1
        for col, lbl in enumerate(["Company", "Period", "Line Item", "Flag Message"], start=1):
            hc = ws.cell(fd_hdr, col, lbl)
            hc.font = HDR_FONT
            hc.fill = HDR_FILL
        fd_row = fd_hdr + 1
        max_msg_len = 0
        for fe in flag_entries:
            ws.cell(fd_row, 1, fe["company"]).font = BASE_FONT
            ws.cell(fd_row, 2, fe["period"]).font  = BASE_FONT
            ws.cell(fd_row, 3, fe["li"]).font      = BASE_FONT
            mc = ws.cell(fd_row, 4, fe["msg"])
            mc.font = Font(name=BF, color="856404", size=11)
            max_msg_len = max(max_msg_len, len(fe["msg"]))
            fd_row += 1
        ws.column_dimensions["D"].width = min(80, max_msg_len + 4)

    # ---- Sheet 2: Source Tags ----
    # One line per value rather than per line item. Per-period resolution means
    # a company's row can span several tags, and a single "tag used" cell would
    # be describing only whichever period happened to come last.
    ws2 = wb.create_sheet("Source Tags")
    ws2.sheet_properties.tabColor = "888888"

    NOTE_FONT = Font(name=BF, color="666666", size=10)
    SEAM_FONT = Font(name=BF, color="856404", size=10)

    tag_headers = [
        ("Company", 30), ("Line Item", 28), ("Period", 10), ("Source", 11),
        ("XBRL Tag", 46), ("Filed", 12), ("Accession", 22), ("Notes", 80),
    ]
    for ci, (text, width) in enumerate(tag_headers, start=1):
        hc = ws2.cell(1, ci, text)
        hc.font = HDR_FONT
        hc.fill = HDR_FILL
        ws2.column_dimensions[hc.column_letter].width = width

    tag_row = 2
    for company in companies:
        for li in line_items:
            item_info = company["line_items"].get(li) or {}
            for period in item_info.get("periods", []):
                prov = period.get("provenance") or {}
                state = prov.get("state")
                if state is None:
                    # A caller that predates provenance still gets an honest
                    # sheet: value present means reported, absent means missing.
                    state = (xbrl.PROVENANCE_REPORTED if period.get("value") is not None
                             else xbrl.PROVENANCE_MISSING)

                note = ""
                if state == xbrl.PROVENANCE_DERIVED:
                    note = "Derived: {}".format(prov.get("formula", ""))
                elif state == xbrl.PROVENANCE_MISSING:
                    note = prov.get("message", "")
                else:
                    seam = next((f for f in period.get("flags", [])
                                 if f.get("flag_type") == xbrl.FLAG_TAG_TRANSITION), None)
                    if seam:
                        note = seam.get("message", "")

                tag = prov.get("tag") or period.get("source_tag") or ""
                # "not found" is a statement about a search. A derived value was
                # never looked for under a tag, so it does not get told one was
                # missing; the Notes column carries its formula instead.
                no_tag = "— arithmetic" if state == xbrl.PROVENANCE_DERIVED else "— not found"
                ws2.cell(tag_row, 1, company["name"]).font = BASE_FONT
                ws2.cell(tag_row, 2, li).font = BASE_FONT
                ws2.cell(tag_row, 3, period.get("relative_period", "")).font = BASE_FONT
                ws2.cell(tag_row, 4, state).font = (
                    MISS_FONT if state == xbrl.PROVENANCE_MISSING else BASE_FONT)
                ws2.cell(tag_row, 5, tag or no_tag).font = (
                    TAG_FONT if tag else MISS_FONT)
                ws2.cell(tag_row, 6, prov.get("filed") or "").font = TAG_FONT
                ws2.cell(tag_row, 7, prov.get("accession") or "").font = TAG_FONT
                ws2.cell(tag_row, 8, note).font = SEAM_FONT if "switches XBRL tag" in note else NOTE_FONT
                tag_row += 1

    ws2.freeze_panes = "A2"
    ws2.cell(tag_row + 1, 1,
             "One line per value. reported = the filer tagged it; derived = Edgardly "
             "computed it from reported values, formula in Notes; missing = nobody "
             "tagged it, Notes says where to look."
             ).font = Font(name=BF, color="666666", italic=True, size=10)

    _peer_add_chart_sheet(wb, comparison_result)
    wb.save(filepath)


@app.route("/api/xbrl/peer-comparison/stream")
def api_peer_comparison_stream():
    """SSE endpoint: fetches one company at a time, yielding progress events,
    then a final 'result' event containing the full comparison data + scale.

    Query params:
        cik (repeated)           -- CIKs to include
        n_periods                -- fiscal years per company (default 3)
        line_items (repeated)    -- which line items (default: all TAG_MAP keys)
    """
    ciks = request.args.getlist("cik")
    if not ciks:
        def _err():
            yield 'data: ' + json.dumps({"type": "error", "message": "cik parameter required"}) + '\n\n'
        return Response(stream_with_context(_err()), mimetype="text/event-stream")

    try:
        n_periods = max(1, min(10, int(request.args.get("n_periods", "3"))))
    except (ValueError, TypeError):
        n_periods = 3

    req_items = request.args.getlist("line_items")
    line_items = req_items if req_items else list(xbrl.TAG_MAP.keys())

    def generate():
        companies = []
        total = len(ciks)
        for i, cik in enumerate(ciks):
            yield 'data: ' + json.dumps({
                "type": "progress", "current": i, "total": total, "cik": cik,
            }) + '\n\n'
            try:
                company_data = pc.fetch_peer_data(str(cik), line_items, n_periods)
                companies.append(company_data)
            except Exception as exc:
                yield 'data: ' + json.dumps({
                    "type": "company_error", "cik": cik, "message": str(exc),
                }) + '\n\n'

        result = {"companies": companies, "line_items": line_items, "n_periods": n_periods}
        result["scale"] = pc.select_peer_scale(result)
        yield 'data: ' + json.dumps({"type": "result", "data": result}) + '\n\n'

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/xbrl/peer-comparison/chart-data", methods=["POST"])
def api_peer_chart_data():
    """Transform an already-fetched comparison_result into chart-ready series.

    Accepts the same comparison_result JSON the SSE stream emits so the
    frontend can call this immediately after the stream completes without
    triggering additional EDGAR requests.
    """
    data = request.get_json(silent=True) or {}
    comparison_result = data.get("comparison_result")
    if not comparison_result or not comparison_result.get("companies"):
        return jsonify({"error": "comparison_result with companies is required"}), 400
    try:
        return jsonify(_build_peer_chart_data(comparison_result))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/xbrl/peer-comparison/export", methods=["POST"])
def api_peer_comparison_export():
    """Write a peer comparison result to an Excel file and return a download URL."""
    data = request.get_json(silent=True) or {}
    comparison_result = data.get("comparison_result")
    if not comparison_result or not comparison_result.get("companies"):
        return jsonify({"error": "comparison_result with companies is required"}), 400

    try:
        company_names = [c.get("name", "Unknown") for c in comparison_result["companies"]]
        safe_names = "_vs_".join(
            "".join(ch if ch.isalnum() or ch in " -_" else "" for ch in n)[:20].strip().replace(" ", "_")
            for n in company_names
        )
        if not safe_names:
            safe_names = "Peer_Comparison"
        folder = os.path.join(EXPORTS_DIR, "Peer_Comparisons")
        os.makedirs(folder, exist_ok=True)
        now      = datetime.datetime.now()
        filename = "Peer_{}_{}_{}.xlsx".format(
            safe_names, now.strftime("%Y-%m-%d"), now.strftime("%H%M")
        )
        filepath = os.path.join(folder, filename)
        _peer_write_xlsx(filepath, comparison_result)
        rel_path = os.path.relpath(filepath, EXPORTS_DIR).replace("\\", "/")
        return jsonify({
            "status": "ok",
            "filename": filename,
            "folder": folder,
            "download_url": "/exports/" + rel_path,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


def _ensure_playwright_chromium():
    """Auto-install Playwright's Chromium browser on first run if not present."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            if os.path.exists(p.chromium.executable_path):
                return
    except ImportError:
        print("WARNING: playwright not installed — PDF downloads unavailable.")
        print("  Run: pip install playwright && playwright install chromium")
        return
    except Exception:
        pass  # Fall through to install attempt

    print("\nFirst-time setup: downloading PDF rendering engine (~150 MB)...")
    print("This is a one-time download. Please wait...\n")
    try:
        proc = subprocess.Popen(
            [sys.executable, '-m', 'playwright', 'install', 'chromium'],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        for line in iter(proc.stdout.readline, ''):
            print(line, end='', flush=True)
        proc.wait()
        if proc.returncode != 0:
            raise RuntimeError(f"installer exited with code {proc.returncode}")
        print("\nSetup complete — PDF rendering engine ready.\n")
    except Exception as exc:
        print(f"\nWARNING: Auto-install failed: {exc}")
        print("To enable PDF downloads, run manually:")
        print("  playwright install chromium")
        print("HTML downloads will still work.\n")


def open_browser():
    webbrowser.open("http://localhost:5050")


if __name__ == "__main__":
    _ensure_playwright_chromium()
    threading.Timer(1.0, open_browser).start()
    app.run(debug=False, port=5050)
