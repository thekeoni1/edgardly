"""excel.py -- the workbook writer, built as a kit.

This module knows openpyxl and knows no finance. Every number it writes comes
from a model spec built by three_statement.py, and every formula it writes is a
rendering of an expression that spec handed it. Phases 3 and 4 reuse the
primitives -- a statement block, a named input, a forecast row, a check row, a
provenance comment -- for a comps sheet and a DCF, which is why they are
functions with explicit arguments rather than steps inside one exporter.

Four rules run through all of it.

**A formula is written only where the model has a value.** A historical cell
the spec could not compute stays empty rather than carrying arithmetic that
would read the hole as a zero: gross profit with no cost of revenue would come
out equal to revenue, which is worse than blank because it looks like an
answer. The cell is styled as missing and its comment says which filing to
open.

**Reported values are static; everything else is a formula.** A number a filer
tagged is written as that number, blue, with its tag, filing date and accession
in the cell comment. A derived value is the live arithmetic over the cells it
came from, black, so the workbook shows its own working. Total Debt is a
formula over Short-Term Debt, which is itself a formula over three tagged
lines, and every leaf of that descent is a blue cell with a filing behind it.

**A blank assumption produces a blank forecast.** Each forecast year has a
readiness cell that is true only when that year's inputs and every earlier
year's are filled, and every forecast cell in the column is guarded by it. An
untaken branch of an IF never propagates, so a workbook with an empty
Assumptions sheet computes to empty forecasts rather than to zeros or errors.

**Nothing about the shape is hardcoded.** Sheet names, the label column, the
first data column and the header rows all come from a Layout, and the number of
columns comes from the spec. A caller that wants a different arrangement passes
a different Layout rather than editing this file.

**Presentation never changes a number.** The workbook is formatted to the
conventions a model is read by -- millions in the number format, indented
detail under ruled-off subtotals, and a colour per kind of cell -- and every one
of those is a property of how a cell is displayed rather than of what it holds.
A dollar cell stores the whole dollars the filer tagged and shows millions,
which is what keeps a provenance comment true, a check exact to the dollar and a
cell a reader copies out worth what the filing says it is. The one place a scale
reaches a formula is the currency assumption, which is typed in millions and
multiplied back where it is consumed, and it is written down here because it is
the exception.
"""

from collections import namedtuple

import line_items
from scaffold import three_statement as ts


# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------

SheetNames = namedtuple(
    "SheetNames", "assumptions income_statement balance_sheet cash_flow schedules "
                  "checks source_tags")

DEFAULT_SHEETS = SheetNames(
    assumptions="Assumptions",
    income_statement="Income Statement",
    balance_sheet="Balance Sheet",
    cash_flow="Cash Flow",
    schedules="Schedules",
    checks="Checks",
    source_tags="Source Tags",
)

Layout = namedtuple(
    "Layout", "sheets label_column first_data_column title_row header_row first_data_row")

DEFAULT_LAYOUT = Layout(sheets=DEFAULT_SHEETS, label_column=1, first_data_column=2,
                        title_row=1, header_row=3, first_data_row=4)

# Which sheet each statement's rows live on. Keyed by the statement codes the
# registry uses, so a new statement is a new entry here and nowhere else.
_STATEMENT_SHEET = {
    line_items.STATEMENT_IS: "income_statement",
    line_items.STATEMENT_BS: "balance_sheet",
    line_items.STATEMENT_CF: "cash_flow",
}


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

BASE_FONT = "Aptos Narrow"

# Blue for a value the filer reported, black for one this workbook computes,
# green for one it reads off another sheet. The first two are older than v2 and
# the rest of the app already uses them; what changed in Phase 1 is that they
# now mean what they say, and what changed in 6c is the third, which is the
# standard convention's one colour about where a number comes from rather than
# what kind of number it is.
COLOR_REPORTED = "0066CC"
COLOR_DERIVED = "000000"
COLOR_LINK = "008000"
COLOR_MISSING = "AAAAAA"
COLOR_HEADER = "003366"
COLOR_NOTE = "666666"
COLOR_FLAG = "856404"

FILL_HEADER = "003366"
FILL_INPUT = "FFF2CC"
FILL_FLAG = "FFF3CD"
FILL_GREEN = "D4EDDA"
FILL_RED = "F8D7DA"

# Every format ends the positive section with an underscore and a bracket, which
# reserves the width of the closing bracket a negative spends, so a column of
# numbers has its digits in one place whatever their signs. The trailing pair of
# commas in the scaled formats divides the display by a million and leaves the
# stored value alone: the cell still holds the dollars the filer tagged, which
# is what keeps a provenance comment true and a check exact.
FORMAT_DOLLAR = "#,##0_);(#,##0)"
FORMAT_DOLLAR_MILLIONS = "#,##0,,_);(#,##0,,)"
FORMAT_EPS = "0.00_);(0.00)"
FORMAT_SHARES_MILLIONS = "#,##0,,_);(#,##0,,)"
FORMAT_PERCENT = "0.0%"
FORMAT_DAYS = "0.0"

_UNIT_FORMAT = {
    line_items.UNIT_DOLLAR: FORMAT_DOLLAR_MILLIONS,
    line_items.UNIT_EPS: FORMAT_EPS,
    line_items.UNIT_SHARES: FORMAT_SHARES_MILLIONS,
}

# What one unit of a currency assumption is worth in the units the statements
# are stored in. The input cell is labelled and formatted in millions so the
# analyst types 500 and means 500 million, and every formula that consumes the
# assumption multiplies it back, so the arithmetic downstream is in raw dollars
# like the rest of the workbook. Percent and days have no scale to get wrong.
ASSUMPTION_SCALE = {"currency": 1_000_000}

# The currency input holds millions already, so its format must not scale again.
_ASSUMPTION_FORMAT = {"percent": FORMAT_PERCENT, "days": FORMAT_DAYS,
                      "currency": FORMAT_DOLLAR}

# What each sheet says about its own scale, in the header region under the
# title. The Checks sheet is the exception the millions decision names: a check
# exists to show a residual, and a residual of a few thousand dollars shown in
# millions is a displayed zero.
UNITS_NOTE_STATEMENT = ("$ in millions, except per share amounts. Share counts in "
                        "millions. Every cell stores the whole units the filer "
                        "reported; the scale is in the number format only.")
UNITS_NOTE_SCHEDULES = ("$ in millions. Every cell stores the whole units the filer "
                        "reported; the scale is in the number format only.")
UNITS_NOTE_CHECKS = ("$ in whole dollars, deliberately: a check is a residual, and a "
                     "residual of a few thousand dollars shown in millions is a "
                     "displayed zero. Coverage is a percentage.")
UNITS_NOTE_ASSUMPTIONS = ("Percents as percents and days as days. A dollar assumption "
                          "is typed in millions: 500 means 500 million.")

# A check is green when it is this close to zero. The scale is dollars, and the
# model's arithmetic on reported dollars is exact, so anything above a dollar is
# a real difference rather than floating point.
CHECK_TOLERANCE = 1.0


def _openpyxl():
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:  # pragma: no cover - the app cannot export without it
        raise RuntimeError("openpyxl is required: pip install openpyxl")
    return (openpyxl, Comment, FormulaRule, Alignment, Border, Font, PatternFill,
            Side, get_column_letter, DefinedName)


def _font(color=COLOR_DERIVED, size=11, bold=False, italic=False):
    """Every font in the workbook, so the family and size are decided once."""
    from openpyxl.styles import Font
    return Font(name=BASE_FONT, size=size, bold=bold, italic=italic, color=color)


def _indent(level):
    from openpyxl.styles import Alignment
    return Alignment(indent=level)


def reads_another_sheet(formula):
    """Whether a rendered formula reaches off the sheet it is written on.

    A sheet qualifier is the only thing that puts an exclamation mark in a
    formula this module writes, so this is the whole test. It decides the green
    font rather than a list of the six rows that link, because a list goes stale
    the moment a row is added and the colour would then be lying about the one
    cell a reader most needs it to be right about.
    """
    return "!" in (formula or "")


def _value_font(state, formula=None, italic=False):
    """The colour convention, in one place: green beats blue beats black."""
    if state == ts.CELL_MISSING:
        return _font(color=COLOR_MISSING, italic=True)
    if reads_another_sheet(formula):
        return _font(color=COLOR_LINK, italic=italic)
    if state == ts.CELL_REPORTED:
        return _font(color=COLOR_REPORTED, italic=italic)
    return _font(color=COLOR_DERIVED, italic=italic)


def _rule_off(worksheet, layout, row_number, columns, double=False):
    """A single top border on a subtotal, and a double bottom to close a statement.

    Applied to the label column and every data column, so the rule runs the
    width of the row a reader is reading rather than stopping under the numbers.
    """
    from openpyxl.styles import Border, Side

    border = Border(top=Side(style="thin"),
                    bottom=Side(style="double") if double else None)
    for offset in range(-1, columns):
        column = (layout.label_column if offset < 0
                  else layout.first_data_column + offset)
        worksheet.cell(row=row_number, column=column).border = border


# ---------------------------------------------------------------------------
# Names and references
# ---------------------------------------------------------------------------

def assumption_name(key, year_index):
    """The defined name for one assumption in one forecast year.

    Letters, digits and underscores only, never starting with a digit and never
    shaped like a cell reference, which is what Excel requires of a defined name
    and what a malformed one costs: the repair dialog that risk R4 is about.
    """
    return "asm_{}_y{}".format(key, year_index + 1)


def ready_name(year_index):
    return "asm_{}_y{}".format(ts.READY_KEY, year_index + 1)


Placement = namedtuple("Placement", "sheet row")


def _quote(sheet_name):
    """A sheet name as a formula uses it, quoted when it has to be."""
    if all(ch.isalnum() or ch == "_" for ch in sheet_name):
        return sheet_name
    return "'{}'".format(sheet_name.replace("'", "''"))


def cell_reference(placement, column_letter, from_sheet=None):
    """An A1 reference to a placed row, qualified by sheet when it has to be."""
    local = "{}{}".format(column_letter, placement.row)
    if from_sheet is None or from_sheet == placement.sheet:
        return local
    return "{}!{}".format(_quote(placement.sheet), local)


# ---------------------------------------------------------------------------
# Rendering an expression
# ---------------------------------------------------------------------------

class FormulaContext(object):
    """Everything the renderer needs to turn an expression into A1 text.

    placements maps a model row name to where it was written; columns maps a
    period index to a column letter. Both are built once the sheets are laid
    out, which is why rendering is a separate pass from placing.
    """

    def __init__(self, placements, columns, sheet, column_index, year_index=None):
        self.placements = placements
        self.columns = columns
        self.sheet = sheet
        self.column_index = column_index
        self.year_index = year_index

    def at(self, column_index):
        return FormulaContext(self.placements, self.columns, self.sheet,
                              column_index, self.year_index)


def render(expr, context):
    """Render one expression tree to an Excel formula body, without the '='.

    Returns None when the expression reaches a row this workbook did not place
    or a column that does not exist, which is how a reference off the left edge
    of the history stops a formula from being written at all rather than
    becoming a reference to whatever happens to be there.
    """
    kind = expr[0]

    if kind == "const":
        return _number(expr[1])

    if kind == "asm":
        if context.year_index is None:
            return None
        name = assumption_name(expr[1], context.year_index)
        # The input cell holds the units the analyst types; the formula puts it
        # back on the workbook's own scale before anything else uses it.
        scale = ASSUMPTION_SCALE.get(ts.ASSUMPTIONS_BY_KEY[expr[1]].unit)
        return name if scale is None else "({}*{})".format(name, _number(scale))

    if kind == "ref":
        _tag, row_name, offset = expr
        placement = context.placements.get(row_name)
        column = context.column_index + offset
        if placement is None or column not in context.columns:
            return None
        return cell_reference(placement, context.columns[column], context.sheet)

    parts = [render(part, context) for part in expr[1:]]
    if any(part is None for part in parts):
        return None

    if kind == "+":
        return "({})".format("+".join(parts))
    if kind == "-":
        return "({})".format("-".join(parts))
    if kind == "*":
        return "({})".format("*".join(parts))
    if kind == "/":
        return "({}/{})".format(parts[0], parts[1])
    raise ValueError("unknown expression node {!r}".format(kind))


def render_terms(terms, context):
    """Render (row, sign, offset) terms as a signed sum of cell references."""
    parts = []
    for row_name, sign, offset in terms:
        rendered = render(("ref", row_name, offset), context)
        if rendered is None:
            return None
        if not parts:
            parts.append(rendered if sign > 0 else "-{}".format(rendered))
        else:
            parts.append("{}{}".format("+" if sign > 0 else "-", rendered))
    return "".join(parts) or None


def _number(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return repr(value) if isinstance(value, float) else str(value)


def guarded(body, year_index):
    """Wrap a forecast body so an unready assumption column produces a blank.

    Excel returns only the branch it takes and never propagates the other one,
    so a column whose neighbours are blank strings still evaluates cleanly to a
    blank rather than to an error. That is what makes "delete the assumptions
    and every forecast cell empties" a property of the workbook rather than a
    hope.
    """
    return "=IF({},{},\"\")".format(ready_name(year_index), body)


# ---------------------------------------------------------------------------
# Primitives
# ---------------------------------------------------------------------------

def attach_provenance_comment(cell, provenance, extra_flags=(), author="Edgardly"):
    """Put a value's provenance in the cell comment, in the reader's language.

    A reported value names its tag, the filing that carried it and when it was
    filed. A derived one carries its formula and each input, and where an input
    is itself derived it carries that input's own formula too, so the descent
    is readable without leaving the cell. A missing one carries the sentence
    that says which statement of which filing to open.
    """
    from openpyxl.comments import Comment

    lines = _provenance_lines(provenance)
    for flag in extra_flags:
        lines.append("[{}] {}".format(flag["flag_type"], flag["message"]))
    if not lines:
        return None
    comment = Comment("\n".join(lines), author)
    comment.width = 380
    comment.height = 30 + 14 * min(len(" ".join(lines)) // 55 + len(lines), 24)
    cell.comment = comment
    return comment


def _provenance_lines(provenance, depth=0):
    if not provenance:
        return []
    indent = "    " * depth
    state = provenance.get("state")
    lines = []
    if state == ts.CELL_REPORTED:
        lines.append("{}Reported by the filer.".format(indent))
        lines.append("{}Tag: {}".format(indent, provenance.get("tag")))
        lines.append("{}Filed: {}  Form: {}".format(
            indent, provenance.get("filed"), provenance.get("form")))
        lines.append("{}Accession: {}".format(indent, provenance.get("accession")))
    elif state == ts.CELL_DERIVED:
        lines.append("{}Computed by Edgardly.".format(indent))
        lines.append("{}Formula: {}".format(indent, provenance.get("formula")))
        for entry in provenance.get("inputs", []):
            lines.append("{}  {} = {}".format(indent, entry.get("name"),
                                              _format_input(entry.get("value"))))
            if entry.get("tag"):
                lines.append("{}      tag {}, filed {}, accession {}".format(
                    indent, entry.get("tag"), entry.get("filed"),
                    entry.get("accession")))
            elif entry.get("formula"):
                lines.append("{}      itself computed as {}".format(
                    indent, entry.get("formula")))
                for nested in entry.get("inputs", []):
                    lines.append("{}        {} = {} (tag {})".format(
                        indent, nested.get("name"),
                        _format_input(nested.get("value")), nested.get("tag")))
    elif state == ts.CELL_MISSING:
        lines.append("{}Not available.".format(indent))
        lines.append("{}{}".format(indent, provenance.get("message", "")))
    return lines


def _format_input(value):
    if value is None:
        return "not reported"
    return "{:,.0f}".format(value) if abs(value) >= 1000 else "{:,.4g}".format(value)


def define_named_input(workbook, worksheet, name, coordinate, note=""):
    """Create a blank, styled input cell and give it a defined name.

    Named one at a time and tested, which is what risk R4 asks for: a malformed
    defined name is one of the few things that makes Excel offer to repair a
    file, and repairing strips content.
    """
    from openpyxl.styles import PatternFill
    from openpyxl.workbook.defined_name import DefinedName

    cell = worksheet[coordinate]
    cell.value = None
    cell.font = _font(color=COLOR_REPORTED)
    cell.fill = PatternFill("solid", fgColor=FILL_INPUT)
    attr = "{}!${}${}".format(_quote(worksheet.title),
                              cell.column_letter, cell.row)
    workbook.defined_names.add(DefinedName(name, attr_text=attr))
    if note:
        from openpyxl.comments import Comment
        cell.comment = Comment(note, "Edgardly")
    return cell


def write_statement_block(worksheet, layout, spec, rows, placements, columns,
                          start_row):
    """Write one statement's rows, historicals and forecasts, and place them.

    Returns the next free row. Called once per statement, and by Phase 3 or 4
    for whatever block they need, which is why it takes its rows rather than
    reading them off the spec itself.

    A detail row is indented under the subtotal it belongs to with openpyxl's
    own indent rather than with spaces in front of the label, so the label a
    reader sees is indented and the label a tooltip, a test or a downstream read
    gets is the string the model holds.
    """
    historical = ts.historical_periods(spec)
    forecast = ts.forecast_periods(spec)
    row_number = start_row
    width = len(historical) + len(forecast)
    bottom_line = final_total_row(rows)

    for row in rows:
        label_cell = worksheet.cell(row=row_number, column=layout.label_column)
        label_cell.value = row.label
        label_cell.font = _font(bold=row.role == ts.ROLE_SUBTOTAL,
                                italic=row.role in (ts.ROLE_PLUG, ts.ROLE_MEMO))
        label_cell.alignment = _indent(0 if row.role == ts.ROLE_SUBTOTAL else 1)
        _attach_row_note(label_cell, row)

        placement = placements[row.name]
        for index, period in enumerate(historical):
            column = layout.first_data_column + index
            _write_historical_cell(worksheet, row, period, row_number, column,
                                   placements, columns, layout)
        for offset, _period in enumerate(forecast):
            index = len(historical) + offset
            column = layout.first_data_column + index
            _write_forecast_cell(worksheet, row, row_number, column, index, offset,
                                 placements, columns, layout)
        if row.role == ts.ROLE_SUBTOTAL or row.name == bottom_line:
            _rule_off(worksheet, layout, row_number, width,
                      double=row.name == bottom_line)
        row_number += 1
    return row_number


def final_total_row(rows):
    """The row a statement ends on, which is the one that gets the double rule.

    The last row that is a subtotal or a derivation, so the memo lines a
    statement carries after its bottom line -- earnings per share, share counts,
    EBITDA, total debt -- do not take the rule off it. Read off the rows rather
    than named per statement, so a block Phase 3 or 4 passes in is closed by the
    same rule this one is.
    """
    for row in reversed(list(rows)):
        if row.role in (ts.ROLE_SUBTOTAL, ts.ROLE_DERIVED):
            return row.name
    return None


def _attach_row_note(cell, row):
    """The row's own note: what the chain does not say, and why it forecasts."""
    from openpyxl.comments import Comment

    lines = []
    if row.note:
        lines.append(row.note)
    if row.forecast_note:
        lines.append("Forecast: {}".format(row.forecast_note))
    for flag in row.flags:
        lines.append("[{}] {}".format(flag["flag_type"], flag["message"]))
    if not lines:
        return
    comment = Comment("\n\n".join(lines), "Edgardly")
    comment.width = 420
    comment.height = 30 + 13 * min(len(" ".join(lines)) // 60 + 2 * len(lines), 30)
    cell.comment = comment


def _write_historical_cell(worksheet, row, period, row_number, column, placements,
                           columns, layout):
    """One historical cell: the filer's number, this workbook's arithmetic, or a hole."""
    from openpyxl.styles import PatternFill

    cell = worksheet.cell(row=row_number, column=column)
    model_cell = row.cells[period.key]
    number_format = _UNIT_FORMAT.get(row.unit, FORMAT_DOLLAR_MILLIONS)

    if model_cell.state == ts.CELL_REPORTED:
        cell.value = model_cell.value
        cell.font = _value_font(ts.CELL_REPORTED)
    elif model_cell.state == ts.CELL_DERIVED:
        context = FormulaContext(placements, columns, worksheet.title,
                                 column - layout.first_data_column)
        body = render_terms(model_cell.terms, context)
        # No formula is better than one that reads a hole as a zero, and the
        # spec has already decided this cell has a value, so a body that cannot
        # be rendered means a reference off the edge of the history rather than
        # a disagreement about the number.
        cell.value = "={}".format(body) if body else model_cell.value
        cell.font = _value_font(ts.CELL_DERIVED, body,
                                italic=row.role == ts.ROLE_PLUG)
    else:
        cell.value = None
        cell.font = _value_font(ts.CELL_MISSING)

    cell.number_format = number_format
    if model_cell.flags:
        cell.fill = PatternFill("solid", fgColor=FILL_FLAG)
    attach_provenance_comment(cell, model_cell.provenance, model_cell.flags)
    return cell


def write_forecast_formula(worksheet, expr, row_number, column, column_index,
                           year_index, placements, columns, sheet_title,
                           number_format=FORMAT_DOLLAR_MILLIONS):
    """One forecast cell: guarded arithmetic, or nothing at all.

    Nothing at all is a real answer here. A row with no assumption behind it and
    no anchor to carry forward gets an empty cell, and an empty cell contributes
    zero to the sums above it, which is right for a line the filer's statements
    do not carry and honest for one this tool cannot model.
    """
    cell = worksheet.cell(row=row_number, column=column)
    context = FormulaContext(placements, columns, sheet_title, column_index,
                             year_index)
    body = render(expr, context) if expr is not None else None
    if body is None:
        cell.value = None
        cell.font = _value_font(ts.CELL_MISSING)
        return cell
    cell.value = guarded(body, year_index)
    cell.font = _value_font(ts.CELL_DERIVED, body)
    cell.number_format = number_format
    return cell


def _write_forecast_cell(worksheet, row, row_number, column, column_index,
                         year_index, placements, columns, layout):
    return write_forecast_formula(
        worksheet, row.forecast, row_number, column, column_index, year_index,
        placements, columns, worksheet.title,
        _UNIT_FORMAT.get(row.unit, FORMAT_DOLLAR_MILLIONS))


def write_check_row(worksheet, layout, spec, check, placements, columns, row_number):
    """One tie-out row: live arithmetic, and green or red on whether it holds.

    A check that this filer's data disarmed is written without the colouring and
    with the reason in its comment. Colouring a zero green when the zero was
    constructed rather than found is the one thing a checks sheet must not do.

    These are the rows that keep the unscaled dollar format. Everything else in
    the workbook displays in millions, and a residual of a few thousand dollars
    displayed in millions is a zero, which would make a broken check look like a
    passing one.
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    historical = ts.historical_periods(spec)
    forecast = ts.forecast_periods(spec)

    label = worksheet.cell(row=row_number, column=layout.label_column)
    label.value = check["name"]
    label.font = _font(bold=True)
    _attach_check_note(label, check)

    written = []
    for index, period in enumerate(historical):
        column = layout.first_data_column + index
        cell = worksheet.cell(row=row_number, column=column)
        cell.number_format = FORMAT_DOLLAR
        model_cell = check["cells"][period.key]
        if model_cell.value is None:
            cell.font = _value_font(ts.CELL_MISSING)
            continue
        context = FormulaContext(placements, columns, worksheet.title, index)
        body = render_terms(check["terms"], context)
        if body is None:
            cell.font = _value_font(ts.CELL_MISSING)
            continue
        cell.value = "={}".format(body)
        cell.font = _value_font(ts.CELL_DERIVED, body)
        written.append(cell.coordinate)

    for offset, _period in enumerate(forecast):
        index = len(historical) + offset
        column = layout.first_data_column + index
        context = FormulaContext(placements, columns, worksheet.title, index, offset)
        body = render_terms(check["terms"], context)
        cell = worksheet.cell(row=row_number, column=column)
        cell.number_format = FORMAT_DOLLAR
        if body is None:
            continue
        cell.value = guarded(body, offset)
        cell.font = _value_font(ts.CELL_DERIVED, body)
        written.append(cell.coordinate)

    if check["tie"] and written:
        first = layout.first_data_column
        last = first + len(historical) + len(forecast) - 1
        span = "{}{}:{}{}".format(get_column_letter(first), row_number,
                                  get_column_letter(last), row_number)
        anchor = "{}{}".format(get_column_letter(first), row_number)
        worksheet.conditional_formatting.add(span, FormulaRule(
            formula=["AND(ISNUMBER({0}),ABS({0})<{1})".format(anchor, CHECK_TOLERANCE)],
            fill=PatternFill("solid", fgColor=FILL_GREEN), stopIfTrue=False))
        worksheet.conditional_formatting.add(span, FormulaRule(
            formula=["AND(ISNUMBER({0}),ABS({0})>={1})".format(anchor,
                                                               CHECK_TOLERANCE)],
            fill=PatternFill("solid", fgColor=FILL_RED), stopIfTrue=False))
    return row_number + 1


def write_coverage_row(worksheet, layout, spec, entry, placements, columns,
                       row_number):
    """One section's coverage: the components' share of the subtotal, live.

    Written as (subtotal less plug) divided by subtotal, which is arithmetic on
    two cells already in the workbook rather than a percentage this module
    computed and typed in. A reader who distrusts the figure can follow it to
    the same two cells the statement shows.

    Only where both cells hold a number and the subtotal is not zero. A ratio
    over an empty cell is a division by zero, and Excel showing #DIV/0! across
    a Checks sheet is the repair-dialog kind of damage in slower motion.
    """
    label = worksheet.cell(row=row_number, column=layout.label_column)
    label.value = entry["name"]
    label.font = _font()
    label.alignment = _indent(1)

    for index, period in enumerate(ts.historical_periods(spec)):
        column = layout.first_data_column + index
        cell = worksheet.cell(row=row_number, column=column)
        cell.number_format = FORMAT_PERCENT
        if entry["cells"].get(period.key) is None:
            cell.font = _value_font(ts.CELL_MISSING)
            continue
        context = FormulaContext(placements, columns, worksheet.title, index)
        total = render(ts.ref(entry["total_row"]), context)
        plug = render(ts.ref(entry["plug_row"]), context)
        if total is None or plug is None:
            cell.font = _value_font(ts.CELL_MISSING)
            continue
        cell.value = "=({}-{})/{}".format(total, plug, total)
        cell.font = _value_font(ts.CELL_DERIVED, total)
    return row_number + 1


def _attach_check_note(cell, check):
    from openpyxl.comments import Comment

    lines = [check["note"]]
    for flag in check["flags"]:
        lines.append("[{}] {}".format(flag["flag_type"], flag["message"]))
    comment = Comment("\n\n".join(lines), "Edgardly")
    comment.width = 440
    comment.height = 200
    cell.comment = comment


# ---------------------------------------------------------------------------
# Composing the workbook
# ---------------------------------------------------------------------------

def _period_columns(layout, spec):
    """Period index to column letter, for every column the model has."""
    from openpyxl.utils import get_column_letter
    return {index: get_column_letter(layout.first_data_column + index)
            for index in range(len(spec.periods))}


def _placements(layout, spec):
    """Where every model row will be written, decided before anything is."""
    placements = {}
    row_number = layout.first_data_row
    for code, _title, _block in ts.STATEMENT_BLOCKS:
        sheet = getattr(layout.sheets, _STATEMENT_SHEET[code])
        row_number = layout.first_data_row
        for row in ts.rows_for(spec, code):
            placements[row.name] = Placement(sheet, row_number)
            row_number += 1
    return placements


def _write_header(worksheet, layout, spec, subtitle, units_note):
    """Title, scale note and column headings, in the rows the layout reserves.

    The scale note goes between the title and the column headings, which is
    where a reader of any model looks for it, and it is written only if the
    layout leaves a row there. It says what the number format does and what the
    cell still holds, because those are now two different things and a reader
    who copies a cell out needs to know which one they are getting.
    """
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    title = worksheet.cell(row=layout.title_row, column=layout.label_column)
    title.value = "{} -- {}".format(spec.entity, subtitle)
    title.font = _font(size=12, bold=True, color=COLOR_HEADER)

    if units_note and layout.title_row + 1 < layout.header_row:
        note = worksheet.cell(row=layout.title_row + 1, column=layout.label_column)
        note.value = units_note
        note.font = _font(size=10, italic=True, color=COLOR_NOTE)

    header = worksheet.cell(row=layout.header_row, column=layout.label_column)
    header.value = "Line item"
    header.font = _font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor=FILL_HEADER)

    for index, period in enumerate(spec.periods):
        cell = worksheet.cell(row=layout.header_row,
                              column=layout.first_data_column + index)
        cell.value = period.label
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.alignment = Alignment(horizontal="right")
        worksheet.column_dimensions[
            get_column_letter(layout.first_data_column + index)].width = 16
    worksheet.column_dimensions[get_column_letter(layout.label_column)].width = 58
    worksheet.freeze_panes = worksheet.cell(row=layout.first_data_row,
                                            column=layout.first_data_column)


def _write_assumptions(workbook, layout, spec):
    """The blank sheet the analyst fills, and the readiness row that reads it."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName

    worksheet = workbook.create_sheet(layout.sheets.assumptions)
    forecast = ts.forecast_periods(spec)

    title = worksheet.cell(row=layout.title_row, column=layout.label_column)
    title.value = "{} -- assumptions. Every cell is yours to fill.".format(spec.entity)
    title.font = _font(size=12, bold=True, color=COLOR_HEADER)

    if layout.title_row + 1 < layout.header_row:
        note = worksheet.cell(row=layout.title_row + 1, column=layout.label_column)
        note.value = UNITS_NOTE_ASSUMPTIONS
        note.font = _font(size=10, italic=True, color=COLOR_NOTE)

    header = worksheet.cell(row=layout.header_row, column=layout.label_column)
    header.value = "Assumption"
    header.font = _font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor=FILL_HEADER)
    for index, period in enumerate(forecast):
        cell = worksheet.cell(row=layout.header_row,
                              column=layout.first_data_column + index)
        cell.value = period.label
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.alignment = Alignment(horizontal="right")
        worksheet.column_dimensions[
            get_column_letter(layout.first_data_column + index)].width = 16
    worksheet.column_dimensions[get_column_letter(layout.label_column)].width = 48

    row_number = layout.first_data_row
    for assumption in spec.assumptions:
        label = worksheet.cell(row=row_number, column=layout.label_column)
        label.value = assumption_label(assumption)
        label.font = _font()
        label.alignment = _indent(1)
        if assumption.note:
            from openpyxl.comments import Comment
            label.comment = Comment(assumption.note, "Edgardly")
        for index, _period in enumerate(forecast):
            column = layout.first_data_column + index
            coordinate = "{}{}".format(get_column_letter(column), row_number)
            cell = define_named_input(workbook, worksheet,
                                      assumption_name(assumption.key, index),
                                      coordinate, assumption.note)
            cell.number_format = _ASSUMPTION_FORMAT.get(assumption.unit,
                                                        FORMAT_DOLLAR)
        row_number += 1

    row_number += 1
    label = worksheet.cell(row=row_number, column=layout.label_column)
    label.value = "Year ready to model?"
    label.font = _font(bold=True)
    from openpyxl.comments import Comment
    label.comment = Comment(
        "True only when every assumption for this year and every earlier year is "
        "filled in. Every forecast cell in the workbook is guarded by this, so a "
        "half-filled column produces no forecast at all rather than a model built "
        "partly on numbers and partly on nothing. Clear the assumptions and the "
        "whole forecast empties.", "Edgardly")

    for index, _period in enumerate(forecast):
        column = layout.first_data_column + index
        cell = worksheet.cell(row=row_number, column=column)
        checks = ["ISBLANK({})".format(assumption_name(a.key, index))
                  for a in spec.assumptions]
        body = "IF(OR({}),FALSE,TRUE)".format(",".join(checks))
        if index:
            body = "IF({},{},FALSE)".format(ready_name(index - 1), body)
        cell.value = "={}".format(body)
        cell.font = _font(bold=True)
        attr = "{}!${}${}".format(_quote(worksheet.title),
                                  get_column_letter(column), row_number)
        workbook.defined_names.add(DefinedName(ready_name(index), attr_text=attr))
    _rule_off(worksheet, layout, row_number, len(forecast))

    return worksheet


def assumption_label(assumption):
    """The label on the input row, which has to say what units to type.

    An assumption in dollars is typed in millions, and a cell that does not say
    so beside statements displaying millions is how a model acquires an
    order-of-magnitude error nobody sees. The model spec's label is untouched;
    this is what the sheet shows.
    """
    if assumption.unit in ASSUMPTION_SCALE:
        return "{} ($ in millions)".format(assumption.label)
    return assumption.label


def _write_statements(workbook, layout, spec, placements, columns):
    written = {}
    for code, subtitle, _block in ts.STATEMENT_BLOCKS:
        worksheet = workbook.create_sheet(getattr(layout.sheets,
                                                  _STATEMENT_SHEET[code]))
        _write_header(worksheet, layout, spec, subtitle, UNITS_NOTE_STATEMENT)
        write_statement_block(worksheet, layout, spec, ts.rows_for(spec, code),
                              placements, columns, layout.first_data_row)
        written[code] = worksheet
    return written


# The roll-forwards, presented on their own sheet. Every line points at the
# statement cell that already holds it rather than restating the arithmetic, so
# there is one place each of these is decided and a schedule cannot drift away
# from the statements it summarises.
#
# Historically a roll-forward does not foot, and pretending otherwise would be
# the same mistake plug rows exist to avoid. Opening property, plant and
# equipment plus capex less depreciation is not the closing balance for any real
# filer, because impairments, disposals, acquisitions and currency move it too
# and none of those is a registry item. So each schedule carries its own
# residual line, named for what actually sits in it, and the residual is zero in
# the forecast columns where the roll-forward is the definition of the row.
Schedule = namedtuple("Schedule", "title opening movements closing residual extras")

_SCHEDULES = (
    Schedule("Property, plant and equipment",
             opening=("Opening balance", "PP&E Net"),
             movements=(("Plus capital expenditure", "Capex", 1),
                        ("Less depreciation and amortisation", "D&A", -1)),
             closing=("Closing balance", "PP&E Net"),
             residual="Impairments, disposals, acquisitions and currency (implied)",
             extras=()),
    Schedule("Debt",
             opening=("Opening long-term debt", "Long-Term Debt"),
             movements=(),
             closing=("Closing long-term debt", "Long-Term Debt"),
             residual="Net issuance or repayment in the year (implied)",
             extras=(("Closing short-term debt", "Short-Term Debt"),
                     ("Closing total debt", "Total Debt"))),
    Schedule("Retained earnings",
             opening=("Opening balance", "Retained Earnings"),
             movements=(("Plus net income", "Net Income", 1),
                        ("Less dividends paid", "Dividends Paid", -1)),
             closing=("Closing balance", "Retained Earnings"),
             residual="Share retirements and other equity movements (implied)",
             extras=()),
    Schedule("Cash",
             opening=("Opening balance", "Cash and Equivalents"),
             movements=(("Plus cash from operations", "Cash from Operations", 1),
                        ("Plus cash from investing", "Cash from Investing", 1),
                        ("Plus cash from financing", "Cash from Financing", 1)),
             closing=("Closing balance", "Cash and Equivalents"),
             residual="Effect of exchange rates on cash and other (implied)",
             extras=()),
)


def _has_value(spec, row_name, period_key):
    row = ts.row_named(spec, row_name)
    if row is None:
        return False
    cell = row.cells.get(period_key)
    return cell is not None and cell.value is not None


def _write_schedule_line(worksheet, layout, spec, label, terms, placements, columns,
                         row_number, closing=False):
    """One line of a schedule: signed references to cells the statements hold.

    A historical column whose sources are not all reported is left empty rather
    than shown as a zero, which is the same rule the statements themselves
    follow and the reason the Schedules sheet cannot invent an opening balance
    for the first year of the model.

    Every figure here is a reference to a statement sheet, so every written cell
    is green: on this sheet the convention says the whole page is a reading of
    the statements rather than a second place any of it is decided.
    """
    historical = ts.historical_periods(spec)
    forecast = ts.forecast_periods(spec)

    cell = worksheet.cell(row=row_number, column=layout.label_column)
    cell.value = label
    cell.font = _font(bold=closing)
    cell.alignment = _indent(0 if closing else 1)
    if closing:
        _rule_off(worksheet, layout, row_number, len(historical) + len(forecast))

    for index in range(len(historical) + len(forecast)):
        column = layout.first_data_column + index
        forecast_index = index - len(historical)
        year_index = forecast_index if forecast_index >= 0 else None
        target = worksheet.cell(row=row_number, column=column)
        target.number_format = FORMAT_DOLLAR_MILLIONS
        if year_index is None:
            known = True
            for name, _sign, offset in terms:
                position = index + offset
                if position < 0 or not _has_value(spec, name,
                                                  historical[position].key):
                    known = False
                    break
            if not known:
                continue
        context = FormulaContext(placements, columns, worksheet.title, index,
                                 year_index)
        body = render_terms(terms, context)
        if body is None:
            continue
        target.value = ("={}".format(body) if year_index is None
                        else guarded(body, year_index))
        target.font = _value_font(ts.CELL_DERIVED, body)
    return row_number + 1


def _write_schedules(workbook, layout, spec, placements, columns):
    worksheet = workbook.create_sheet(layout.sheets.schedules)
    _write_header(worksheet, layout, spec, "Schedules", UNITS_NOTE_SCHEDULES)

    row_number = layout.first_data_row
    for schedule in _SCHEDULES:
        heading = worksheet.cell(row=row_number, column=layout.label_column)
        heading.value = schedule.title
        heading.font = _font(bold=True, color=COLOR_HEADER)
        row_number += 1

        opening_label, opening_row = schedule.opening
        opening_terms = ((opening_row, 1, -1),)
        row_number = _write_schedule_line(worksheet, layout, spec, opening_label,
                                          opening_terms, placements, columns,
                                          row_number)
        for label, source, sign in schedule.movements:
            row_number = _write_schedule_line(worksheet, layout, spec, label,
                                              ((source, sign, 0),), placements,
                                              columns, row_number)

        closing_label, closing_row = schedule.closing
        residual_terms = (((closing_row, 1, 0), (opening_row, -1, -1))
                          + tuple((source, -sign, 0)
                                  for _label, source, sign in schedule.movements))
        row_number = _write_schedule_line(worksheet, layout, spec, schedule.residual,
                                          residual_terms, placements, columns,
                                          row_number)
        row_number = _write_schedule_line(worksheet, layout, spec, closing_label,
                                          ((closing_row, 1, 0),), placements,
                                          columns, row_number, closing=True)
        for label, source in schedule.extras:
            row_number = _write_schedule_line(worksheet, layout, spec, label,
                                              ((source, 1, 0),), placements,
                                              columns, row_number)
        row_number += 1
    return worksheet


def _write_checks(workbook, layout, spec, placements, columns):
    worksheet = workbook.create_sheet(layout.sheets.checks)
    _write_header(worksheet, layout, spec, "Checks", UNITS_NOTE_CHECKS)
    row_number = layout.first_data_row
    for check in spec.checks:
        row_number = write_check_row(worksheet, layout, spec, check, placements,
                                     columns, row_number)

    row_number += 1
    note = worksheet.cell(row=row_number, column=layout.label_column)
    note.value = ("Green is within one unit of zero. A row that is not a tie is not "
                  "coloured; hover its label for why. These rows are in whole "
                  "dollars, not millions, so a small residual stays visible.")
    note.font = _font(size=10, italic=True, color=COLOR_NOTE)

    if spec.coverage:
        row_number += 2
        heading = worksheet.cell(row=row_number, column=layout.label_column)
        heading.value = "How much of each balance sheet section this scaffold reaches"
        heading.font = _font(bold=True, color=COLOR_HEADER)
        from openpyxl.comments import Comment
        comment = Comment(ts.COVERAGE_NOTE, "Edgardly")
        comment.width = 440
        comment.height = 180
        heading.comment = comment
        row_number += 1
        for entry in spec.coverage:
            row_number = write_coverage_row(worksheet, layout, spec, entry,
                                            placements, columns, row_number)
        note = worksheet.cell(row=row_number, column=layout.label_column)
        note.value = ("The rest of each subtotal is in its plug row. These are "
                      "measurements, not warnings: hover the heading for what a "
                      "low figure does and does not mean.")
        note.font = _font(size=10, italic=True, color=COLOR_NOTE)
        row_number += 1

    row_number += 2
    heading = worksheet.cell(row=row_number, column=layout.label_column)
    heading.value = "What this scaffold flags about this filer"
    heading.font = _font(bold=True, color=COLOR_HEADER)
    row_number += 1
    for flag in ts.summarised_flags(spec):
        cell = worksheet.cell(row=row_number, column=layout.label_column)
        cell.value = "[{}] {}".format(flag["flag_type"], flag["message"])
        cell.font = _font(size=10, color=COLOR_FLAG)
        row_number += 1
    return worksheet


def _write_source_tags(workbook, layout, spec):
    """One line per historical value, so a reader can trace any cell to a filing."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(layout.sheets.source_tags)
    headings = ("Statement", "Line item", "Period", "State", "Tag", "Filed",
                "Accession", "Note")
    for index, heading in enumerate(headings):
        cell = worksheet.cell(row=1, column=1 + index)
        cell.value = heading
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.alignment = Alignment(horizontal="left")
    for index, width in enumerate((10, 56, 10, 10, 46, 12, 22, 90)):
        worksheet.column_dimensions[get_column_letter(1 + index)].width = width

    row_number = 2
    for period in ts.historical_periods(spec):
        for row in spec.rows:
            model_cell = row.cells[period.key]
            provenance = model_cell.provenance or {}
            note = ""
            if model_cell.state == ts.CELL_DERIVED:
                note = provenance.get("formula", "")
            elif model_cell.state == ts.CELL_MISSING:
                note = provenance.get("message", "")
            if model_cell.flags:
                note = "{} {}".format(
                    note, " ".join("[{}] {}".format(f["flag_type"], f["message"])
                                   for f in model_cell.flags)).strip()
            values = (row.statement, row.name, period.label, model_cell.state,
                      provenance.get("tag") or "", provenance.get("filed") or "",
                      provenance.get("accession") or "", note)
            for index, value in enumerate(values):
                cell = worksheet.cell(row=row_number, column=1 + index)
                cell.value = value
                cell.font = _font(size=10,
                                  color=COLOR_MISSING
                                  if model_cell.state == ts.CELL_MISSING
                                  else COLOR_DERIVED)
            row_number += 1
    worksheet.freeze_panes = "A2"
    return worksheet


def build_workbook(spec, layout=DEFAULT_LAYOUT):
    """The whole workbook for one model spec.

    Raises ValueError for a spec the scope gate refused, because there is no
    honest workbook to build for one; the caller shows scope.message instead.
    """
    openpyxl = _openpyxl()[0]

    if not spec.scope.in_scope:
        raise ValueError(spec.scope.message)
    if not spec.rows:
        raise ValueError("This filer has no fiscal year Edgardly can confirm, so "
                         "there is no model to build.")

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    columns = _period_columns(layout, spec)
    placements = _placements(layout, spec)

    _write_assumptions(workbook, layout, spec)
    _write_statements(workbook, layout, spec, placements, columns)
    _write_schedules(workbook, layout, spec, placements, columns)
    _write_checks(workbook, layout, spec, placements, columns)
    _write_source_tags(workbook, layout, spec)
    return workbook


def write_workbook(spec, path, layout=DEFAULT_LAYOUT):
    workbook = build_workbook(spec, layout)
    workbook.save(path)
    return path
